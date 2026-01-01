import flet as ft
import sqlite3
import datetime
import os
import pathlib
import platform
import hashlib
import json

# Importar openpyxl solo si está disponible (no funciona en Android)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False
    print("openpyxl no disponible - exportación Excel desactivada")

# --- Lógica de Base de Datos (SQLite) ---
class Database:
    def __init__(self, db_path="finanzas.db"):
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self.create_table()

    def create_table(self):
        cursor = self.conn.cursor()
        # Tabla de configuración de la app
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS configuracion (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                clave TEXT UNIQUE NOT NULL,
                valor TEXT
            )
        """)
        # Tabla de presupuestos por categoría
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS presupuestos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                categoria TEXT UNIQUE NOT NULL,
                limite REAL NOT NULL,
                mes INTEGER,
                anio INTEGER
            )
        """)
        # Tabla de movimientos (solo personal)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS movimientos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tipo TEXT,          -- 'ingreso' o 'gasto'
                categoria TEXT,
                monto REAL,
                descripcion TEXT,
                fecha TEXT
            )
        """)
        # Tabla de suscripciones
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS suscripciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                monto REAL NOT NULL,
                dia_cobro INTEGER,  -- día del mes (1-31)
                activa INTEGER DEFAULT 1  -- 1=activa, 0=inactiva
            )
        """)
        # Tabla de préstamos
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS prestamos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                banco TEXT NOT NULL,
                monto_total REAL NOT NULL,
                monto_pagado REAL DEFAULT 0,
                cuota_mensual REAL NOT NULL,
                dia_pago INTEGER,  -- día del mes (1-31)
                fecha_inicio TEXT,
                activo INTEGER DEFAULT 1  -- 1=activo, 0=pagado
            )
        """)
        # Tabla de ahorros
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ahorros (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                meta REAL NOT NULL,
                monto_actual REAL DEFAULT 0,
                fecha_inicio TEXT,
                completado INTEGER DEFAULT 0  -- 1=completado, 0=en progreso
            )
        """)
        # Tabla de compras a crédito
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS creditos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                descripcion TEXT NOT NULL,
                banco TEXT NOT NULL,
                monto_total REAL NOT NULL,
                meses_sin_intereses INTEGER NOT NULL,
                cuota_mensual REAL NOT NULL,
                meses_pagados INTEGER DEFAULT 0,
                fecha_compra TEXT,
                tasa_interes REAL DEFAULT 0,  -- tasa de interés mensual (0 = sin intereses)
                pagado INTEGER DEFAULT 0  -- 1=pagado, 0=en proceso
            )
        """)
        # Tabla de cuentas bancarias
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS cuentas_bancarias (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre_banco TEXT NOT NULL,
                tipo_cuenta TEXT NOT NULL,  -- 'debito', 'credito', 'ahorro', 'inversion'
                saldo REAL DEFAULT 0,
                limite_credito REAL DEFAULT 0,  -- para tarjetas de crédito
                fecha_creacion TEXT,
                activa INTEGER DEFAULT 1  -- 1=activa, 0=inactiva
            )
        """)
        # Tabla de transferencias entre cuentas
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS transferencias (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cuenta_origen INTEGER,
                cuenta_destino INTEGER,
                monto REAL NOT NULL,
                fecha TEXT,
                descripcion TEXT,
                FOREIGN KEY (cuenta_origen) REFERENCES cuentas_bancarias(id),
                FOREIGN KEY (cuenta_destino) REFERENCES cuentas_bancarias(id)
            )
        """)
        self.conn.commit()
    
    # --- Métodos de Configuración ---
    
    def obtener_config(self, clave, default=None):
        """Obtiene un valor de configuración"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT valor FROM configuracion WHERE clave = ?", (clave,))
            result = cursor.fetchone()
            return result[0] if result else default
        except:
            return default
    
    def guardar_config(self, clave, valor):
        """Guarda un valor de configuración"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?, ?)", (clave, valor))
            self.conn.commit()
            return True
        except:
            return False
    
    def verificar_pin(self, pin):
        """Verifica si el PIN es correcto"""
        pin_guardado = self.obtener_config("pin_hash")
        if not pin_guardado:
            return True  # No hay PIN configurado
        pin_hash = hashlib.sha256(pin.encode()).hexdigest()
        return pin_hash == pin_guardado
    
    def guardar_pin(self, pin):
        """Guarda el PIN encriptado"""
        pin_hash = hashlib.sha256(pin.encode()).hexdigest()
        return self.guardar_config("pin_hash", pin_hash)
    
    def tiene_pin(self):
        """Verifica si hay PIN configurado"""
        return self.obtener_config("pin_hash") is not None
    
    def es_primera_vez(self):
        """Verifica si es la primera vez que se abre la app"""
        return self.obtener_config("onboarding_completado") != "1"
    
    def completar_onboarding(self):
        """Marca el onboarding como completado"""
        return self.guardar_config("onboarding_completado", "1")
    
    def obtener_tema(self):
        """Obtiene el tema actual (light/dark)"""
        return self.obtener_config("tema", "light")
    
    def guardar_tema(self, tema):
        """Guarda el tema seleccionado"""
        return self.guardar_config("tema", tema)
    
    # --- Métodos de Presupuestos ---
    
    def agregar_presupuesto(self, categoria, limite):
        """Agrega o actualiza un presupuesto por categoría"""
        try:
            cursor = self.conn.cursor()
            ahora = datetime.datetime.now()
            cursor.execute("""
                INSERT OR REPLACE INTO presupuestos (categoria, limite, mes, anio) 
                VALUES (?, ?, ?, ?)
            """, (categoria, limite, ahora.month, ahora.year))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al agregar presupuesto: {e}")
            return False
    
    def obtener_presupuestos(self):
        """Obtiene todos los presupuestos"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT * FROM presupuestos ORDER BY categoria")
            return cursor.fetchall()
        except:
            return []
    
    def obtener_gasto_categoria_mes(self, categoria, mes=None, anio=None):
        """Obtiene el total gastado en una categoría en un mes"""
        try:
            if mes is None:
                mes = datetime.datetime.now().month
            if anio is None:
                anio = datetime.datetime.now().year
            
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT SUM(monto) FROM movimientos 
                WHERE tipo = 'gasto' AND categoria = ? 
                AND strftime('%m', fecha) = ? AND strftime('%Y', fecha) = ?
            """, (categoria, f"{mes:02d}", str(anio)))
            result = cursor.fetchone()[0]
            return result if result else 0
        except:
            return 0
    
    def borrar_presupuesto(self, id_presupuesto):
        """Elimina un presupuesto"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("DELETE FROM presupuestos WHERE id = ?", (id_presupuesto,))
            self.conn.commit()
            return True
        except:
            return False
    
    # --- Métodos de Transferencias ---
    
    def realizar_transferencia(self, cuenta_origen, cuenta_destino, monto, descripcion=""):
        """Realiza una transferencia entre cuentas"""
        try:
            cursor = self.conn.cursor()
            fecha = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            
            # Retirar de cuenta origen
            self.retirar_monto_cuenta(cuenta_origen, monto)
            # Depositar en cuenta destino
            self.agregar_monto_cuenta(cuenta_destino, monto)
            
            # Registrar la transferencia
            cursor.execute("""
                INSERT INTO transferencias (cuenta_origen, cuenta_destino, monto, fecha, descripcion)
                VALUES (?, ?, ?, ?, ?)
            """, (cuenta_origen, cuenta_destino, monto, fecha, descripcion))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error en transferencia: {e}")
            return False
    
    def obtener_transferencias(self):
        """Obtiene el historial de transferencias"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT t.id, c1.nombre_banco, c2.nombre_banco, t.monto, t.fecha, t.descripcion
                FROM transferencias t
                LEFT JOIN cuentas_bancarias c1 ON t.cuenta_origen = c1.id
                LEFT JOIN cuentas_bancarias c2 ON t.cuenta_destino = c2.id
                ORDER BY t.fecha DESC
            """)
            return cursor.fetchall()
        except:
            return []
    
    # --- Métodos de Estadísticas para Gráficos ---
    
    def obtener_gastos_por_categoria(self, mes=None, anio=None):
        """Obtiene gastos agrupados por categoría"""
        try:
            if mes is None:
                mes = datetime.datetime.now().month
            if anio is None:
                anio = datetime.datetime.now().year
            
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT categoria, SUM(monto) as total
                FROM movimientos 
                WHERE tipo = 'gasto' 
                AND strftime('%m', fecha) = ? AND strftime('%Y', fecha) = ?
                GROUP BY categoria
                ORDER BY total DESC
            """, (f"{mes:02d}", str(anio)))
            return cursor.fetchall()
        except:
            return []
    
    def obtener_balance_ultimos_meses(self, num_meses=6):
        """Obtiene ingresos y gastos de los últimos N meses"""
        try:
            resultados = []
            ahora = datetime.datetime.now()
            
            for i in range(num_meses - 1, -1, -1):
                fecha = ahora - datetime.timedelta(days=i*30)
                mes = fecha.month
                anio = fecha.year
                
                ingresos, gastos = self.obtener_balance_mensual(mes, anio)
                resultados.append({
                    "mes": fecha.strftime("%b"),
                    "anio": anio,
                    "ingresos": ingresos,
                    "gastos": gastos
                })
            
            return resultados
        except:
            return []
    
    # --- Métodos de Edición ---
    
    def editar_movimiento(self, id_mov, tipo, categoria, monto, descripcion):
        """Edita un movimiento existente"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                UPDATE movimientos SET tipo = ?, categoria = ?, monto = ?, descripcion = ?
                WHERE id = ?
            """, (tipo, categoria, monto, descripcion, id_mov))
            self.conn.commit()
            return True
        except:
            return False
    
    def editar_suscripcion(self, id_sub, nombre, monto, dia_cobro):
        """Edita una suscripción existente"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                UPDATE suscripciones SET nombre = ?, monto = ?, dia_cobro = ?
                WHERE id = ?
            """, (nombre, monto, dia_cobro, id_sub))
            self.conn.commit()
            return True
        except:
            return False
    
    def editar_prestamo(self, id_pres, banco, monto_total, cuota_mensual, dia_pago):
        """Edita un préstamo existente"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                UPDATE prestamos SET banco = ?, monto_total = ?, cuota_mensual = ?, dia_pago = ?
                WHERE id = ?
            """, (banco, monto_total, cuota_mensual, dia_pago, id_pres))
            self.conn.commit()
            return True
        except:
            return False
    
    def editar_ahorro(self, id_aho, nombre, meta):
        """Edita un ahorro existente"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                UPDATE ahorros SET nombre = ?, meta = ?
                WHERE id = ?
            """, (nombre, meta, id_aho))
            self.conn.commit()
            return True
        except:
            return False
    
    def editar_credito(self, id_cred, descripcion, banco, monto_total, meses_plazo, tasa_interes):
        """Edita un crédito existente"""
        try:
            cursor = self.conn.cursor()
            if tasa_interes > 0:
                tasa_mensual = tasa_interes / 100
                cuota_mensual = monto_total * (tasa_mensual * pow(1 + tasa_mensual, meses_plazo)) / (pow(1 + tasa_mensual, meses_plazo) - 1)
            else:
                cuota_mensual = monto_total / meses_plazo if meses_plazo > 0 else monto_total
            
            cursor.execute("""
                UPDATE creditos SET descripcion = ?, banco = ?, monto_total = ?, 
                meses_sin_intereses = ?, cuota_mensual = ?, tasa_interes = ?
                WHERE id = ?
            """, (descripcion, banco, monto_total, meses_plazo, cuota_mensual, tasa_interes, id_cred))
            self.conn.commit()
            return True
        except:
            return False
    
    def editar_cuenta_bancaria(self, id_cuenta, nombre_banco, tipo_cuenta, limite_credito):
        """Edita una cuenta bancaria existente"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                UPDATE cuentas_bancarias SET nombre_banco = ?, tipo_cuenta = ?, limite_credito = ?
                WHERE id = ?
            """, (nombre_banco, tipo_cuenta, limite_credito, id_cuenta))
            self.conn.commit()
            return True
        except:
            return False
    
    # --- Métodos de Búsqueda ---
    
    def buscar_movimientos(self, texto="", categoria=None, tipo=None, fecha_desde=None, fecha_hasta=None):
        """Busca movimientos con filtros"""
        try:
            cursor = self.conn.cursor()
            query = "SELECT * FROM movimientos WHERE 1=1"
            params = []
            
            if texto:
                query += " AND (descripcion LIKE ? OR categoria LIKE ?)"
                params.extend([f"%{texto}%", f"%{texto}%"])
            
            if categoria:
                query += " AND categoria = ?"
                params.append(categoria)
            
            if tipo:
                query += " AND tipo = ?"
                params.append(tipo)
            
            if fecha_desde:
                query += " AND fecha >= ?"
                params.append(fecha_desde)
            
            if fecha_hasta:
                query += " AND fecha <= ?"
                params.append(fecha_hasta)
            
            query += " ORDER BY id DESC"
            cursor.execute(query, params)
            return cursor.fetchall()
        except:
            return []
    
    # --- Backup y Restauración ---
    
    def exportar_datos(self):
        """Exporta todos los datos a un diccionario"""
        try:
            cursor = self.conn.cursor()
            datos = {}
            
            tablas = ['movimientos', 'suscripciones', 'prestamos', 'ahorros', 'creditos', 'cuentas_bancarias', 'presupuestos']
            
            for tabla in tablas:
                cursor.execute(f"SELECT * FROM {tabla}")
                columnas = [description[0] for description in cursor.description]
                filas = cursor.fetchall()
                datos[tabla] = [dict(zip(columnas, fila)) for fila in filas]
            
            return json.dumps(datos, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error al exportar: {e}")
            return None
    
    def importar_datos(self, json_data):
        """Importa datos desde un JSON"""
        try:
            datos = json.loads(json_data)
            cursor = self.conn.cursor()
            
            for tabla, registros in datos.items():
                for registro in registros:
                    columnas = ', '.join(registro.keys())
                    placeholders = ', '.join(['?' for _ in registro])
                    valores = list(registro.values())
                    
                    try:
                        cursor.execute(f"INSERT OR REPLACE INTO {tabla} ({columnas}) VALUES ({placeholders})", valores)
                    except:
                        pass
            
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al importar: {e}")
            return False

    def agregar_movimiento(self, tipo, categoria, monto, descripcion):
        try:
            cursor = self.conn.cursor()
            fecha = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            cursor.execute("INSERT INTO movimientos (tipo, categoria, monto, descripcion, fecha) VALUES (?, ?, ?, ?, ?)",
                           (tipo, categoria, monto, descripcion, fecha))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al agregar movimiento: {e}")
            return False

    def obtener_movimientos(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT * FROM movimientos ORDER BY id DESC")
            return cursor.fetchall()
        except Exception as e:
            print(f"Error al obtener movimientos: {e}")
            return []

    def obtener_balance(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT tipo, monto FROM movimientos")
            datos = cursor.fetchall()
            
            ingresos = sum(d[1] for d in datos if d[0] == 'ingreso')
            gastos = sum(d[1] for d in datos if d[0] == 'gasto')
            return ingresos, gastos, (ingresos - gastos)
        except Exception as e:
            print(f"Error al obtener balance: {e}")
            return 0, 0, 0
    
    def obtener_balance_mensual(self, mes, anio):
        """Obtiene balance de un mes específico"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT tipo, monto FROM movimientos 
                WHERE strftime('%m', fecha) = ? AND strftime('%Y', fecha) = ?
            """, (f"{mes:02d}", str(anio)))
            datos = cursor.fetchall()
            
            ingresos = sum(d[1] for d in datos if d[0] == 'ingreso')
            gastos = sum(d[1] for d in datos if d[0] == 'gasto')
            return ingresos, gastos
        except Exception as e:
            print(f"Error al obtener balance mensual: {e}")
            return 0, 0
    
    def obtener_movimientos_mensuales(self, mes, anio):
        """Obtiene todos los movimientos de un mes específico"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT id, tipo, categoria, monto, descripcion, fecha 
                FROM movimientos 
                WHERE strftime('%m', fecha) = ? AND strftime('%Y', fecha) = ?
                ORDER BY fecha DESC
            """, (f"{mes:02d}", str(anio)))
            return cursor.fetchall()
        except Exception as e:
            print(f"Error al obtener movimientos mensuales: {e}")
            return []
    
    # --- Métodos para Suscripciones ---
    
    def agregar_suscripcion(self, nombre, monto, dia_cobro):
        try:
            cursor = self.conn.cursor()
            cursor.execute("INSERT INTO suscripciones (nombre, monto, dia_cobro) VALUES (?, ?, ?)",
                           (nombre, monto, dia_cobro))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al agregar suscripción: {e}")
            return False
    
    def obtener_suscripciones(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT * FROM suscripciones WHERE activa = 1 ORDER BY dia_cobro")
            return cursor.fetchall()
        except Exception as e:
            print(f"Error al obtener suscripciones: {e}")
            return []
    
    def obtener_total_suscripciones(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT SUM(monto) FROM suscripciones WHERE activa = 1")
            result = cursor.fetchone()[0]
            return result if result else 0
        except Exception as e:
            print(f"Error al obtener total de suscripciones: {e}")
            return 0
    
    def borrar_suscripcion(self, id_suscripcion):
        try:
            cursor = self.conn.cursor()
            cursor.execute("UPDATE suscripciones SET activa = 0 WHERE id = ?", (id_suscripcion,))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al borrar suscripción: {e}")
            return False

    def borrar_movimiento(self, id_movimiento):
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM movimientos WHERE id = ?", (id_movimiento,))
        self.conn.commit()
    
    # --- Métodos para Préstamos ---
    
    def agregar_prestamo(self, banco, monto_total, cuota_mensual, dia_pago):
        try:
            cursor = self.conn.cursor()
            fecha_inicio = datetime.datetime.now().strftime("%Y-%m-%d")
            cursor.execute("INSERT INTO prestamos (banco, monto_total, cuota_mensual, dia_pago, fecha_inicio) VALUES (?, ?, ?, ?, ?)",
                           (banco, monto_total, cuota_mensual, dia_pago, fecha_inicio))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al agregar préstamo: {e}")
            return False
    
    def obtener_prestamos(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT * FROM prestamos WHERE activo = 1 ORDER BY dia_pago")
            return cursor.fetchall()
        except Exception as e:
            print(f"Error al obtener préstamos: {e}")
            return []
    
    def obtener_total_cuotas_prestamos(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT SUM(cuota_mensual) FROM prestamos WHERE activo = 1")
            result = cursor.fetchone()[0]
            return result if result else 0
        except Exception as e:
            print(f"Error al obtener total de cuotas: {e}")
            return 0
    
    def obtener_deuda_total(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT SUM(monto_total - monto_pagado) FROM prestamos WHERE activo = 1")
            result = cursor.fetchone()[0]
            return result if result else 0
        except Exception as e:
            print(f"Error al obtener deuda total: {e}")
            return 0
    
    def registrar_pago_prestamo(self, id_prestamo, monto_pago):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT monto_total, monto_pagado FROM prestamos WHERE id = ?", (id_prestamo,))
            result = cursor.fetchone()
            if result:
                monto_total, monto_pagado = result
                nuevo_monto_pagado = monto_pagado + monto_pago
                
                # Si se pagó todo, marcar como inactivo
                if nuevo_monto_pagado >= monto_total:
                    cursor.execute("UPDATE prestamos SET monto_pagado = ?, activo = 0 WHERE id = ?", 
                                   (monto_total, id_prestamo))
                else:
                    cursor.execute("UPDATE prestamos SET monto_pagado = ? WHERE id = ?", 
                                   (nuevo_monto_pagado, id_prestamo))
                self.conn.commit()
                return True
        except Exception as e:
            print(f"Error al registrar pago: {e}")
            return False
    
    def borrar_prestamo(self, id_prestamo):
        try:
            cursor = self.conn.cursor()
            cursor.execute("UPDATE prestamos SET activo = 0 WHERE id = ?", (id_prestamo,))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al borrar préstamo: {e}")
            return False
    
    # --- Métodos para Ahorros ---
    
    def agregar_ahorro(self, nombre, meta):
        try:
            cursor = self.conn.cursor()
            fecha_inicio = datetime.datetime.now().strftime("%Y-%m-%d")
            cursor.execute("INSERT INTO ahorros (nombre, meta, fecha_inicio) VALUES (?, ?, ?)",
                           (nombre, meta, fecha_inicio))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al agregar ahorro: {e}")
            return False
    
    def obtener_ahorros(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT * FROM ahorros WHERE completado = 0 ORDER BY fecha_inicio DESC")
            return cursor.fetchall()
        except Exception as e:
            print(f"Error al obtener ahorros: {e}")
            return []
    
    def obtener_total_ahorros(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT SUM(monto_actual) FROM ahorros WHERE completado = 0")
            result = cursor.fetchone()[0]
            return result if result else 0
        except Exception as e:
            print(f"Error al obtener total de ahorros: {e}")
            return 0
    
    def agregar_monto_ahorro(self, id_ahorro, monto):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT meta, monto_actual FROM ahorros WHERE id = ?", (id_ahorro,))
            result = cursor.fetchone()
            if result:
                meta, monto_actual = result
                nuevo_monto = monto_actual + monto
                
                # Si se alcanzó la meta, marcar como completado
                if nuevo_monto >= meta:
                    cursor.execute("UPDATE ahorros SET monto_actual = ?, completado = 1 WHERE id = ?", 
                                   (meta, id_ahorro))
                else:
                    cursor.execute("UPDATE ahorros SET monto_actual = ? WHERE id = ?", 
                                   (nuevo_monto, id_ahorro))
                self.conn.commit()
                return True
        except Exception as e:
            print(f"Error al agregar monto: {e}")
            return False
    
    def retirar_monto_ahorro(self, id_ahorro, monto):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT monto_actual FROM ahorros WHERE id = ?", (id_ahorro,))
            result = cursor.fetchone()
            if result:
                monto_actual = result[0]
                nuevo_monto = max(0, monto_actual - monto)
                cursor.execute("UPDATE ahorros SET monto_actual = ? WHERE id = ?", 
                               (nuevo_monto, id_ahorro))
                self.conn.commit()
                return True
        except Exception as e:
            print(f"Error al retirar monto: {e}")
            return False
    
    def borrar_ahorro(self, id_ahorro):
        try:
            cursor = self.conn.cursor()
            cursor.execute("UPDATE ahorros SET completado = 1 WHERE id = ?", (id_ahorro,))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al borrar ahorro: {e}")
            return False
    
    # --- Métodos para Compras a Crédito ---
    
    def agregar_credito(self, descripcion, banco, monto_total, meses_plazo, tasa_interes=0):
        try:
            cursor = self.conn.cursor()
            # Si hay interés, calcular cuota con fórmula de amortización
            if tasa_interes > 0:
                tasa_mensual = tasa_interes / 100  # Convertir a decimal
                # Fórmula de cuota fija: P * [r(1+r)^n] / [(1+r)^n - 1]
                cuota_mensual = monto_total * (tasa_mensual * pow(1 + tasa_mensual, meses_plazo)) / (pow(1 + tasa_mensual, meses_plazo) - 1)
            else:
                # Sin intereses, dividir en partes iguales
                cuota_mensual = monto_total / meses_plazo if meses_plazo > 0 else monto_total
            
            fecha_compra = datetime.datetime.now().strftime("%Y-%m-%d")
            cursor.execute("INSERT INTO creditos (descripcion, banco, monto_total, meses_sin_intereses, cuota_mensual, fecha_compra, tasa_interes) VALUES (?, ?, ?, ?, ?, ?, ?)",
                           (descripcion, banco, monto_total, meses_plazo, cuota_mensual, fecha_compra, tasa_interes))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al agregar crédito: {e}")
            return False
    
    def obtener_creditos(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT * FROM creditos WHERE pagado = 0 ORDER BY fecha_compra DESC")
            return cursor.fetchall()
        except Exception as e:
            print(f"Error al obtener créditos: {e}")
            return []
    
    def obtener_total_cuotas_creditos(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT SUM(cuota_mensual) FROM creditos WHERE pagado = 0")
            result = cursor.fetchone()[0]
            return result if result else 0
        except Exception as e:
            print(f"Error al obtener total de cuotas: {e}")
            return 0
    
    def obtener_deuda_total_creditos(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT meses_sin_intereses, meses_pagados, cuota_mensual FROM creditos WHERE pagado = 0")
            creditos = cursor.fetchall()
            total = sum((meses_totales - meses_pagados) * cuota for meses_totales, meses_pagados, cuota in creditos)
            return total
        except Exception as e:
            print(f"Error al obtener deuda total de créditos: {e}")
            return 0
    
    def registrar_pago_credito(self, id_credito):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT meses_sin_intereses, meses_pagados FROM creditos WHERE id = ?", (id_credito,))
            result = cursor.fetchone()
            if result:
                meses_totales, meses_pagados = result
                nuevos_meses_pagados = meses_pagados + 1
                
                # Si se pagaron todos los meses, marcar como pagado
                if nuevos_meses_pagados >= meses_totales:
                    cursor.execute("UPDATE creditos SET meses_pagados = ?, pagado = 1 WHERE id = ?", 
                                   (meses_totales, id_credito))
                else:
                    cursor.execute("UPDATE creditos SET meses_pagados = ? WHERE id = ?", 
                                   (nuevos_meses_pagados, id_credito))
                self.conn.commit()
                return True
        except Exception as e:
            print(f"Error al registrar pago de crédito: {e}")
            return False
    
    def borrar_credito(self, id_credito):
        try:
            cursor = self.conn.cursor()
            cursor.execute("UPDATE creditos SET pagado = 1 WHERE id = ?", (id_credito,))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al borrar crédito: {e}")
            return False
    
    # --- Métodos para Cuentas Bancarias ---
    
    def agregar_cuenta_bancaria(self, nombre_banco, tipo_cuenta, saldo_inicial=0, limite_credito=0):
        try:
            cursor = self.conn.cursor()
            fecha_creacion = datetime.datetime.now().strftime("%Y-%m-%d")
            cursor.execute("INSERT INTO cuentas_bancarias (nombre_banco, tipo_cuenta, saldo, limite_credito, fecha_creacion) VALUES (?, ?, ?, ?, ?)",
                           (nombre_banco, tipo_cuenta, saldo_inicial, limite_credito, fecha_creacion))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al agregar cuenta bancaria: {e}")
            return False
    
    def obtener_cuentas_bancarias(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT * FROM cuentas_bancarias WHERE activa = 1 ORDER BY nombre_banco")
            return cursor.fetchall()
        except Exception as e:
            print(f"Error al obtener cuentas bancarias: {e}")
            return []
    
    def obtener_saldo_total_bancos(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT SUM(saldo) FROM cuentas_bancarias WHERE activa = 1")
            result = cursor.fetchone()[0]
            return result if result else 0
        except Exception as e:
            print(f"Error al obtener saldo total: {e}")
            return 0
    
    def actualizar_saldo_cuenta(self, id_cuenta, nuevo_saldo):
        try:
            cursor = self.conn.cursor()
            cursor.execute("UPDATE cuentas_bancarias SET saldo = ? WHERE id = ?", (nuevo_saldo, id_cuenta))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al actualizar saldo: {e}")
            return False
    
    def agregar_monto_cuenta(self, id_cuenta, monto):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT saldo FROM cuentas_bancarias WHERE id = ?", (id_cuenta,))
            result = cursor.fetchone()
            if result:
                nuevo_saldo = result[0] + monto
                return self.actualizar_saldo_cuenta(id_cuenta, nuevo_saldo)
        except Exception as e:
            print(f"Error al agregar monto: {e}")
            return False
    
    def retirar_monto_cuenta(self, id_cuenta, monto):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT saldo FROM cuentas_bancarias WHERE id = ?", (id_cuenta,))
            result = cursor.fetchone()
            if result:
                nuevo_saldo = result[0] - monto
                return self.actualizar_saldo_cuenta(id_cuenta, nuevo_saldo)
        except Exception as e:
            print(f"Error al retirar monto: {e}")
            return False
    
    def borrar_cuenta_bancaria(self, id_cuenta):
        try:
            cursor = self.conn.cursor()
            cursor.execute("UPDATE cuentas_bancarias SET activa = 0 WHERE id = ?", (id_cuenta,))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error al borrar cuenta: {e}")
            return False
    
    def close(self):
        """Cierra la conexión a la base de datos"""
        if self.conn:
            self.conn.close()


def get_persistent_db_path():
    """
    Obtiene una ruta persistente para la base de datos.
    Esta ruta NO se elimina cuando se actualiza la app.
    """
    try:
        sistema = platform.system().lower()
        
        if sistema == 'windows':
            # Windows: usar AppData/Local
            app_data_dir = pathlib.Path.home() / "AppData" / "Local" / "JFinanzas"
        elif sistema == 'darwin':
            # macOS: usar Application Support
            app_data_dir = pathlib.Path.home() / "Library" / "Application Support" / "JFinanzas"
        elif sistema == 'linux':
            # Verificar si es Android (Flet en Android reporta como Linux)
            # En Android, usar el directorio de datos de la app
            home = pathlib.Path.home()
            if 'data' in str(home) and 'app' in str(home).lower():
                # Es Android - usar directorio actual
                app_data_dir = pathlib.Path(".") / "data"
            else:
                # Linux normal
                app_data_dir = home / ".jfinanzas"
        else:
            # Otros sistemas (iOS, etc)
            app_data_dir = pathlib.Path.home() / ".jfinanzas"
        
        app_data_dir.mkdir(parents=True, exist_ok=True)
        return str(app_data_dir / "finanzas.db")
    except Exception as e:
        print(f"Error obteniendo ruta persistente: {e}")
        # Fallback: usar directorio actual
        try:
            data_dir = pathlib.Path("data")
            data_dir.mkdir(exist_ok=True)
            return str(data_dir / "finanzas.db")
        except:
            return "finanzas.db"


# --- Interfaz Gráfica (Flet) ---
def main(page: ft.Page):
    # Configuración inicial para móvil
    page.title = "💰 Mis Finanzas"
    page.padding = 0
    page.scroll = ft.ScrollMode.AUTO
    
    # Configuración óptima para móvil y PC
    try:
        if hasattr(page, 'web') and not page.web:
            page.window_width = 400
            page.window_height = 800
    except:
        pass
    
    # Usar almacenamiento persistente con manejo de errores
    try:
        db_path = get_persistent_db_path()
        print(f"Base de datos en: {db_path}")
        db = Database(db_path)
    except Exception as e:
        # Si falla, mostrar error en pantalla
        page.add(ft.Text(f"Error al inicializar: {e}", color="red", size=16))
        page.update()
        return
    
    # Aplicar tema guardado
    tema_guardado = db.obtener_tema()
    page.theme_mode = ft.ThemeMode.DARK if tema_guardado == "dark" else ft.ThemeMode.LIGHT
    
    # Colores según tema - Paleta oscura moderna y agradable
    def obtener_colores():
        es_oscuro = page.theme_mode == ft.ThemeMode.DARK
        if es_oscuro:
            return {
                "fondo": "#0d1117",
                "tarjeta": "#161b22",
                "tarjeta_elevada": "#21262d",
                "texto": "#e6edf3",
                "texto_secundario": "#8b949e",
                "borde": "#30363d",
                "appbar": "#161b22",
                "input_border": "#30363d",
                "input_bg": "#0d1117",
                # Colores de acento para modo oscuro
                "verde": "#238636",
                "verde_bg": "#0d1117",
                "rojo": "#da3633",
                "rojo_bg": "#0d1117",
                "naranja": "#d29922",
                "naranja_bg": "#1c1504",
                "purple": "#8957e5",
                "purple_bg": "#1a0d2e",
                "azul": "#58a6ff",
                "azul_bg": "#0d1117",
                "teal": "#3fb950",
                "teal_bg": "#0d1a0f",
                "cyan": "#39c5cf",
                "cyan_bg": "#0a1a1c",
                "indigo": "#a371f7",
                "indigo_bg": "#170d2e",
                "gris_bg": "#21262d",
            }
        else:
            return {
                "fondo": "white",
                "tarjeta": "white",
                "tarjeta_elevada": "#f6f8fa",
                "texto": "black",
                "texto_secundario": "grey600",
                "borde": "#e0e0e0",
                "appbar": "blue700",
                "input_border": "blue700",
                "input_bg": "white",
                # Colores de acento para modo claro
                "verde": "green",
                "verde_bg": "green50",
                "rojo": "red",
                "rojo_bg": "red50",
                "naranja": "orange",
                "naranja_bg": "orange50",
                "purple": "purple",
                "purple_bg": "purple50",
                "azul": "blue700",
                "azul_bg": "blue50",
                "teal": "teal",
                "teal_bg": "teal50",
                "cyan": "cyan900",
                "cyan_bg": "cyan50",
                "indigo": "indigo",
                "indigo_bg": "indigo50",
                "gris_bg": "grey100",
            }
    
    colores = obtener_colores()
    
    # Estado para navegación
    vista_actual = "inicio"
    app_desbloqueada = [False]  # Usar lista para poder modificar en funciones anidadas
    
    # =====================================================
    # PANTALLA DE PIN / ONBOARDING
    # =====================================================
    
    contenedor_login = ft.Container(expand=True, visible=True)
    contenedor_app = ft.Container(expand=True, visible=False)
    
    # Campos para PIN
    pin_inputs = []
    for i in range(4):
        pin_inputs.append(ft.TextField(
            width=50,
            height=60,
            text_align=ft.TextAlign.CENTER,
            keyboard_type=ft.KeyboardType.NUMBER,
            max_length=1,
            text_size=24,
            border_radius=10,
            password=True,
            on_change=lambda e, idx=i: manejar_pin_input(e, idx)
        ))
    
    txt_pin_mensaje = ft.Text("", color="red", size=14, text_align=ft.TextAlign.CENTER)
    txt_pin_titulo = ft.Text("Ingresa tu PIN", size=24, weight=ft.FontWeight.BOLD, text_align=ft.TextAlign.CENTER)
    
    def manejar_pin_input(e, idx):
        """Maneja la entrada de PIN y pasa al siguiente campo"""
        if e.control.value and idx < 3:
            pin_inputs[idx + 1].focus()
        
        # Si es el último dígito, verificar PIN
        if idx == 3 and e.control.value:
            pin_completo = "".join([p.value for p in pin_inputs])
            if len(pin_completo) == 4:
                verificar_pin_ingresado(pin_completo)
    
    def verificar_pin_ingresado(pin):
        """Verifica el PIN ingresado"""
        if db.tiene_pin():
            if db.verificar_pin(pin):
                desbloquear_app()
            else:
                txt_pin_mensaje.value = "❌ PIN incorrecto"
                limpiar_pin()
                page.update()
        else:
            # Es nuevo, guardar PIN
            guardar_nuevo_pin(pin)
    
    def guardar_nuevo_pin(pin):
        """Guarda un nuevo PIN"""
        if db.guardar_pin(pin):
            txt_pin_mensaje.value = "✅ PIN creado correctamente"
            page.update()
            import time
            time.sleep(0.5)
            desbloquear_app()
    
    def limpiar_pin():
        """Limpia los campos de PIN"""
        for p in pin_inputs:
            p.value = ""
        pin_inputs[0].focus()
    
    def desbloquear_app():
        """Desbloquea la app y muestra la pantalla principal"""
        app_desbloqueada[0] = True
        contenedor_login.visible = False
        contenedor_app.visible = True
        
        # Si es primera vez, mostrar onboarding
        if db.es_primera_vez():
            mostrar_onboarding()
        else:
            actualizar_vista()
        
        page.update()
    
    def saltar_pin():
        """Permite saltar el PIN (solo si no hay PIN configurado)"""
        if not db.tiene_pin():
            desbloquear_app()
    
    # Botón para saltar PIN (solo visible si no hay PIN)
    btn_saltar_pin = ft.TextButton(
        "Continuar sin PIN",
        on_click=lambda e: saltar_pin(),
        visible=not db.tiene_pin()
    )
    
    # =====================================================
    # PANTALLA DE ONBOARDING
    # =====================================================
    
    onboarding_index = [0]
    
    onboarding_pages = [
        {
            "icono": "account_balance_wallet",
            "titulo": "¡Bienvenido a Mis Finanzas!",
            "descripcion": "Tu asistente personal para controlar ingresos, gastos y alcanzar tus metas financieras.",
            "color": "blue"
        },
        {
            "icono": "trending_up",
            "titulo": "Controla tus Movimientos",
            "descripcion": "Registra fácilmente todos tus ingresos y gastos. Categoriza y mantén un historial completo.",
            "color": "green"
        },
        {
            "icono": "subscriptions",
            "titulo": "Gestiona Suscripciones",
            "descripcion": "Nunca pierdas de vista tus pagos recurrentes como Netflix, Spotify y más.",
            "color": "orange"
        },
        {
            "icono": "savings",
            "titulo": "Alcanza tus Metas",
            "descripcion": "Crea metas de ahorro y visualiza tu progreso. ¡Cada peso cuenta!",
            "color": "teal"
        },
        {
            "icono": "pie_chart",
            "titulo": "Visualiza tus Finanzas",
            "descripcion": "Gráficos y reportes para entender mejor cómo gastas tu dinero.",
            "color": "purple"
        }
    ]
    
    contenedor_onboarding = ft.Container(visible=False, expand=True)
    
    def crear_pagina_onboarding(index):
        """Crea una página del onboarding"""
        p = onboarding_pages[index]
        return ft.Container(
            content=ft.Column([
                ft.Container(height=50),
                ft.Icon(p["icono"], size=120, color=p["color"]),
                ft.Container(height=30),
                ft.Text(p["titulo"], size=28, weight=ft.FontWeight.BOLD, text_align=ft.TextAlign.CENTER),
                ft.Container(height=20),
                ft.Text(p["descripcion"], size=16, text_align=ft.TextAlign.CENTER, color="grey600"),
                ft.Container(height=50),
                # Indicadores de página
                ft.Row(
                    [ft.Container(
                        width=10 if i != index else 30,
                        height=10,
                        border_radius=5,
                        bgcolor=p["color"] if i == index else "grey300"
                    ) for i in range(len(onboarding_pages))],
                    alignment=ft.MainAxisAlignment.CENTER,
                    spacing=5
                ),
                ft.Container(height=30),
                ft.Row([
                    ft.TextButton("Saltar", on_click=lambda e: finalizar_onboarding()) if index < len(onboarding_pages) - 1 else ft.Container(),
                    ft.ElevatedButton(
                        "Siguiente" if index < len(onboarding_pages) - 1 else "¡Comenzar!",
                        on_click=lambda e: siguiente_onboarding(),
                        bgcolor=p["color"],
                        color="white"
                    )
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            padding=30,
            expand=True
        )
    
    def mostrar_onboarding():
        """Muestra la pantalla de onboarding"""
        onboarding_index[0] = 0
        contenedor_onboarding.content = crear_pagina_onboarding(0)
        contenedor_onboarding.visible = True
        contenedor_app.visible = False
        page.update()
    
    def siguiente_onboarding():
        """Avanza a la siguiente página del onboarding"""
        if onboarding_index[0] < len(onboarding_pages) - 1:
            onboarding_index[0] += 1
            contenedor_onboarding.content = crear_pagina_onboarding(onboarding_index[0])
            page.update()
        else:
            finalizar_onboarding()
    
    def finalizar_onboarding():
        """Finaliza el onboarding y muestra la app"""
        db.completar_onboarding()
        contenedor_onboarding.visible = False
        contenedor_app.visible = True
        actualizar_vista()
        page.update()
    
    # =====================================================
    # COMPONENTES PRINCIPALES DE LA APP
    # =====================================================
    
    # Texto del Balance
    txt_balance_total = ft.Text("$0", size=36, weight=ft.FontWeight.BOLD)
    txt_ingresos = ft.Text("$0", color="green", size=16, weight=ft.FontWeight.BOLD)
    txt_gastos = ft.Text("$0", color="red", size=16, weight=ft.FontWeight.BOLD)
    txt_suscripciones = ft.Text("$0", color="orange", size=16, weight=ft.FontWeight.BOLD)
    txt_prestamos = ft.Text("$0", color="purple", size=16, weight=ft.FontWeight.BOLD)
    txt_creditos = ft.Text("$0", color="indigo", size=16, weight=ft.FontWeight.BOLD)
    txt_ahorros = ft.Text("$0", color="teal", size=16, weight=ft.FontWeight.BOLD)
    txt_bancos = ft.Text("$0", color="cyan900", size=16, weight=ft.FontWeight.BOLD)
    txt_disponible = ft.Text("$0", size=20, weight=ft.FontWeight.BOLD, color="blue700")

    # Contenedores para diferentes vistas
    contenedor_principal = ft.Column(spacing=0, expand=True)

    # Campos para agregar nuevo movimiento
    input_desc = ft.TextField(
        label="Descripción",
        hint_text="Ej: Supermercado",
        color=colores["texto"],
        text_size=16,
        border_color=colores["input_border"],
        focused_border_color="blue900"
    )
    input_monto = ft.TextField(
        label="Monto",
        keyboard_type=ft.KeyboardType.NUMBER,
        color=colores["texto"],
        text_size=16,
        border_color=colores["input_border"],
        focused_border_color="blue900"
    )
    dropdown_tipo = ft.Dropdown(
        label="Tipo",
        options=[ft.dropdown.Option("gasto"), ft.dropdown.Option("ingreso")],
        value="gasto",
        on_change=lambda e: actualizar_opciones_destino()
    )
    dropdown_cat = ft.Dropdown(
        label="Categoría",
        options=[
            ft.dropdown.Option("Comida", "🍔 Comida"),
            ft.dropdown.Option("Transporte", "🚗 Transporte"),
            ft.dropdown.Option("Servicios", "💡 Servicios"),
            ft.dropdown.Option("Ocio", "🎮 Ocio"),
            ft.dropdown.Option("Salud", "💊 Salud"),
            ft.dropdown.Option("Salario", "💰 Salario"),
            ft.dropdown.Option("Compras", "🛒 Compras"),
            ft.dropdown.Option("Educación", "📚 Educación"),
            ft.dropdown.Option("Otro", "📌 Otro"),
        ],
        value="Comida"
    )
    dropdown_destino_movimiento = ft.Dropdown(
        label="Método de pago",
        options=[
            ft.dropdown.Option("efectivo", "💵 Efectivo"),
            ft.dropdown.Option("banco", "🏦 Banco")
        ],
        value="efectivo",
        visible=True,
        on_change=lambda e: actualizar_selector_banco()
    )
    dropdown_banco_movimiento = ft.Dropdown(
        label="Selecciona el banco",
        options=[],
        visible=False
    )
    
    def actualizar_opciones_destino():
        """Actualiza las etiquetas y opciones según el tipo de movimiento"""
        es_ingreso = dropdown_tipo.value == "ingreso"
        
        # Cambiar la etiqueta según el tipo
        if es_ingreso:
            dropdown_destino_movimiento.label = "Destino del ingreso"
        else:
            dropdown_destino_movimiento.label = "Método de pago"
        
        dropdown_banco_movimiento.visible = False
        
        # Obtener bancos disponibles
        cuentas = db.obtener_cuentas_bancarias()
        opciones = []
        for cuenta in cuentas:
            id_cuenta, nombre_banco, tipo_cuenta, saldo, limite_credito, fecha_creacion, activa = cuenta
            opciones.append(ft.dropdown.Option(f"banco_{id_cuenta}", f"{nombre_banco} ({tipo_cuenta})"))
        dropdown_banco_movimiento.options = opciones
        if opciones:
            dropdown_banco_movimiento.value = opciones[0].key
        
        page.update()
    
    def actualizar_selector_banco():
        """Muestra el selector de banco solo si se elige 'banco'"""
        dropdown_banco_movimiento.visible = dropdown_destino_movimiento.value == "banco"
        page.update()

    # --- Funciones de Lógica ---

    def actualizar_balance():
        """Actualiza los textos del balance"""
        ingresos, gastos, total = db.obtener_balance()
        total_suscripciones = db.obtener_total_suscripciones()
        total_cuotas_prestamos = db.obtener_total_cuotas_prestamos()
        total_cuotas_creditos = db.obtener_total_cuotas_creditos()
        total_ahorros = db.obtener_total_ahorros()
        total_bancos = db.obtener_saldo_total_bancos()
        disponible = total - total_suscripciones - total_cuotas_prestamos - total_cuotas_creditos
        
        txt_balance_total.value = f"${total:,.0f}"
        txt_ingresos.value = f"${ingresos:,.0f}"
        txt_gastos.value = f"${gastos:,.0f}"
        txt_suscripciones.value = f"${total_suscripciones:,.0f}"
        txt_prestamos.value = f"${total_cuotas_prestamos:,.0f}"
        txt_creditos.value = f"${total_cuotas_creditos:,.0f}"
        txt_ahorros.value = f"${total_ahorros:,.0f}"
        txt_bancos.value = f"${total_bancos:,.0f}"
        txt_disponible.value = f"${disponible:,.0f}"

    # =====================================================
    # DIÁLOGOS DE CONFIRMACIÓN Y EDICIÓN
    # =====================================================
    
    dialogo_confirmacion = ft.AlertDialog(
        modal=True,
        title=ft.Text("⚠️ Confirmar eliminación"),
        content=ft.Text("¿Estás seguro de que deseas eliminar este registro?"),
        actions=[
            ft.TextButton("Cancelar", on_click=lambda e: cerrar_dialogo()),
            ft.ElevatedButton("Eliminar", bgcolor="red", color="white", on_click=lambda e: None),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )
    
    def cerrar_dialogo():
        dialogo_confirmacion.open = False
        page.update()
    
    def confirmar_borrado(tipo, id_registro, nombre=""):
        """Muestra diálogo de confirmación antes de borrar"""
        dialogo_confirmacion.content = ft.Text(f"¿Eliminar {nombre}?")
        dialogo_confirmacion.actions[1].on_click = lambda e: ejecutar_borrado(tipo, id_registro)
        dialogo_confirmacion.open = True
        page.update()
    
    def ejecutar_borrado(tipo, id_registro):
        """Ejecuta el borrado después de confirmación"""
        if tipo == "movimiento":
            db.borrar_movimiento(id_registro)
        elif tipo == "suscripcion":
            db.borrar_suscripcion(id_registro)
        elif tipo == "prestamo":
            db.borrar_prestamo(id_registro)
        elif tipo == "ahorro":
            db.borrar_ahorro(id_registro)
        elif tipo == "credito":
            db.borrar_credito(id_registro)
        elif tipo == "cuenta":
            db.borrar_cuenta_bancaria(id_registro)
        elif tipo == "presupuesto":
            db.borrar_presupuesto(id_registro)
        
        cerrar_dialogo()
        actualizar_vista()
    
    # Variables para edición
    registro_editando = [None, None]  # [tipo, id]
    
    # =====================================================
    # BÚSQUEDA Y FILTROS
    # =====================================================
    
    input_busqueda = ft.TextField(
        label="🔍 Buscar",
        hint_text="Buscar movimientos...",
        border_radius=20,
        on_change=lambda e: aplicar_filtros()
    )
    
    filtro_categoria = ft.Dropdown(
        label="Categoría",
        options=[ft.dropdown.Option("", "Todas")] + [
            ft.dropdown.Option("Comida"), ft.dropdown.Option("Transporte"),
            ft.dropdown.Option("Servicios"), ft.dropdown.Option("Ocio"),
            ft.dropdown.Option("Salud"), ft.dropdown.Option("Salario"),
            ft.dropdown.Option("Compras"), ft.dropdown.Option("Educación"),
            ft.dropdown.Option("Otro")
        ],
        value="",
        on_change=lambda e: aplicar_filtros()
    )
    
    filtro_tipo = ft.Dropdown(
        label="Tipo",
        options=[
            ft.dropdown.Option("", "Todos"),
            ft.dropdown.Option("ingreso", "Ingresos"),
            ft.dropdown.Option("gasto", "Gastos")
        ],
        value="",
        on_change=lambda e: aplicar_filtros()
    )
    
    mostrar_filtros = [False]
    
    def aplicar_filtros():
        """Aplica los filtros y actualiza la vista"""
        actualizar_vista()
    
    def toggle_filtros():
        """Muestra/oculta los filtros"""
        mostrar_filtros[0] = not mostrar_filtros[0]
        actualizar_vista()
    
    # =====================================================
    # VISTA DE INICIO CON MEJORAS
    # =====================================================
    
    def crear_vista_inicio():
        """Crea la vista principal con gráficos interactivos y movimientos"""
        colores = obtener_colores()
        
        # Datos del mes actual
        ahora = datetime.datetime.now()
        mes_actual = ahora.month
        anio_actual = ahora.year
        meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        mes_nombre = meses_nombres[mes_actual - 1]
        
        # Obtener datos financieros del mes
        ingresos_mes, gastos_mes = db.obtener_balance_mensual(mes_actual, anio_actual)
        total_subs = db.obtener_total_suscripciones()
        total_cuotas = db.obtener_total_cuotas_prestamos()
        total_creditos = db.obtener_total_cuotas_creditos()
        gastos_fijos = total_subs + total_cuotas + total_creditos
        disponible_mes = ingresos_mes - gastos_mes - gastos_fijos
        
        # Gastos por categoría
        gastos_categoria = db.obtener_gastos_por_categoria(mes_actual, anio_actual)
        
        # Colores para categorías
        colores_cat = {
            "Comida": "#FF6384", "Transporte": "#36A2EB", "Servicios": "#FFCE56",
            "Ocio": "#4BC0C0", "Salud": "#9966FF", "Salario": "#2ECC71",
            "Compras": "#FF9F40", "Educación": "#E74C3C", "Otro": "#95A5A6"
        }
        
        # === GRÁFICO CIRCULAR DE DISTRIBUCIÓN ===
        def crear_grafico_circular():
            total_egresos = gastos_mes + gastos_fijos
            if total_egresos <= 0:
                return ft.Container(
                    content=ft.Text("Sin gastos este mes", color=colores["texto_secundario"], italic=True),
                    alignment=ft.alignment.center,
                    padding=20
                )
            
            # Calcular porcentajes para el anillo
            pct_gastos = (gastos_mes / total_egresos * 100) if total_egresos > 0 else 0
            pct_fijos = (gastos_fijos / total_egresos * 100) if total_egresos > 0 else 0
            
            return ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Stack([
                            ft.Container(
                                width=100, height=100,
                                border_radius=50,
                                bgcolor=colores["borde"],
                            ),
                            ft.Container(
                                width=100, height=100,
                                content=ft.Column([
                                    ft.Text(f"${total_egresos:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                                    ft.Text("Total", size=10, color=colores["texto_secundario"])
                                ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                            ),
                        ]),
                        ft.Column([
                            ft.Row([
                                ft.Container(width=12, height=12, bgcolor=colores["rojo"], border_radius=3),
                                ft.Text(f"Gastos: {pct_gastos:.0f}%", size=12, color=colores["texto"])
                            ], spacing=8),
                            ft.Row([
                                ft.Container(width=12, height=12, bgcolor=colores["naranja"], border_radius=3),
                                ft.Text(f"Fijos: {pct_fijos:.0f}%", size=12, color=colores["texto"])
                            ], spacing=8),
                        ], spacing=8)
                    ], alignment=ft.MainAxisAlignment.SPACE_AROUND),
                ]),
                padding=10
            )
        
        # === BARRA DE PROGRESO DEL MES ===
        def crear_barra_progreso_mes():
            dia_actual = ahora.day
            dias_mes = 31 if mes_actual in [1,3,5,7,8,10,12] else 30 if mes_actual in [4,6,9,11] else 29 if anio_actual % 4 == 0 else 28
            progreso_dias = dia_actual / dias_mes
            
            # Calcular si vamos bien o mal
            if ingresos_mes > 0:
                progreso_gastos = (gastos_mes + gastos_fijos) / ingresos_mes
                estado_color = colores["verde"] if progreso_gastos <= progreso_dias else colores["rojo"]
                estado_texto = "✅ Vas bien" if progreso_gastos <= progreso_dias else "⚠️ Cuidado"
            else:
                progreso_gastos = 0
                estado_color = colores["texto_secundario"]
                estado_texto = "Sin ingresos"
            
            return ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Text(f"📅 Día {dia_actual} de {dias_mes}", size=13, color=colores["texto"]),
                        ft.Text(estado_texto, size=13, color=estado_color, weight=ft.FontWeight.BOLD)
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    ft.ProgressBar(value=progreso_dias, color=colores["azul"], bgcolor=colores["borde"], height=8),
                    ft.Row([
                        ft.Text(f"Gastado: ${gastos_mes + gastos_fijos:,.0f}", size=11, color=colores["rojo"]),
                        ft.Text(f"de ${ingresos_mes:,.0f}", size=11, color=colores["texto_secundario"])
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ], spacing=8),
                padding=15,
                bgcolor=colores["tarjeta"],
                border_radius=12,
                margin=ft.margin.only(left=10, right=10, bottom=10)
            )
        
        # === MINI GRÁFICO DE CATEGORÍAS ===
        def crear_grafico_categorias_mini():
            if not gastos_categoria:
                return ft.Container()
            
            total_gastos = sum([g[1] for g in gastos_categoria])
            max_gasto = max([g[1] for g in gastos_categoria]) if gastos_categoria else 1
            
            # Solo mostrar top 4 categorías
            top_cats = sorted(gastos_categoria, key=lambda x: x[1], reverse=True)[:4]
            
            barras = []
            for cat, monto in top_cats:
                pct = (monto / total_gastos * 100) if total_gastos > 0 else 0
                ancho = (monto / max_gasto) if max_gasto > 0 else 0
                color = colores_cat.get(cat, "#95A5A6")
                
                barras.append(
                    ft.Row([
                        ft.Container(
                            content=ft.Text(cat[:8], size=10, color=colores["texto"]),
                            width=65
                        ),
                        ft.Container(
                            content=ft.Container(
                                width=ancho * 120,
                                height=16,
                                bgcolor=color,
                                border_radius=4,
                            ),
                            bgcolor=colores["borde"],
                            border_radius=4,
                            width=120,
                            height=16,
                        ),
                        ft.Text(f"{pct:.0f}%", size=10, color=colores["texto_secundario"], width=35)
                    ], spacing=5)
                )
            
            return ft.Container(
                content=ft.Column([
                    ft.Text("🏷️ Top Gastos", size=13, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                    ft.Divider(height=5, color="transparent"),
                    *barras
                ], spacing=6),
                padding=15,
                bgcolor=colores["tarjeta"],
                border_radius=12,
                margin=ft.margin.only(left=10, right=10, bottom=10)
            )
        
        # === RESUMEN RÁPIDO ===
        def crear_resumen_rapido():
            return ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Text(f"💰 {mes_nombre} {anio_actual}", size=18, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                    ]),
                    ft.Divider(height=10, color="transparent"),
                    ft.Row([
                        ft.Container(
                            content=ft.Column([
                                ft.Icon("trending_up", color=colores["verde"], size=22),
                                ft.Text(f"${ingresos_mes:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["verde"]),
                                ft.Text("Ingresos", size=10, color=colores["texto_secundario"])
                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                            bgcolor=colores["verde_bg"],
                            padding=10,
                            border_radius=10,
                            expand=True
                        ),
                        ft.Container(
                            content=ft.Column([
                                ft.Icon("trending_down", color=colores["rojo"], size=22),
                                ft.Text(f"${gastos_mes:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["rojo"]),
                                ft.Text("Gastos", size=10, color=colores["texto_secundario"])
                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                            bgcolor=colores["rojo_bg"],
                            padding=10,
                            border_radius=10,
                            expand=True
                        ),
                        ft.Container(
                            content=ft.Column([
                                ft.Icon("account_balance_wallet", color=colores["azul"], size=22),
                                ft.Text(f"${disponible_mes:,.0f}", size=14, weight=ft.FontWeight.BOLD, 
                                       color=colores["verde"] if disponible_mes >= 0 else colores["rojo"]),
                                ft.Text("Disponible", size=10, color=colores["texto_secundario"])
                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                            bgcolor=colores["azul_bg"],
                            padding=10,
                            border_radius=10,
                            expand=True
                        ),
                    ], spacing=8),
                ]),
                padding=15,
                margin=10,
                bgcolor=colores["tarjeta"],
                border_radius=15,
                shadow=ft.BoxShadow(spread_radius=0, blur_radius=8, color="black12", offset=ft.Offset(0, 2))
            )
        
        # === LISTA DE MOVIMIENTOS ===
        lista_movimientos = ft.ListView(spacing=8, padding=10, expand=True)
        
        texto_busqueda = input_busqueda.value if input_busqueda.value else ""
        cat_filtro = filtro_categoria.value if filtro_categoria.value else None
        tipo_filtro = filtro_tipo.value if filtro_tipo.value else None
        
        if texto_busqueda or cat_filtro or tipo_filtro:
            movimientos = db.buscar_movimientos(texto_busqueda, cat_filtro, tipo_filtro)
        else:
            movimientos = db.obtener_movimientos()[:10]  # Solo últimos 10
        
        if not movimientos:
            lista_movimientos.controls.append(
                ft.Container(
                    content=ft.Text("No hay movimientos aún.\n¡Agrega tu primer movimiento!", 
                                   italic=True, text_align=ft.TextAlign.CENTER, size=14, color=colores["texto_secundario"]),
                    padding=30
                )
            )
        else:
            for mov in movimientos:
                if len(mov) == 7:
                    id_mov, tipo, cat, monto, desc, fecha, modo = mov
                else:
                    id_mov, tipo, cat, monto, desc, fecha = mov
                
                icono = "trending_down" if tipo == "gasto" else "trending_up"
                color_icono = colores["rojo"] if tipo == "gasto" else colores["verde"]
                
                item = ft.Container(
                    content=ft.Row([
                        ft.Icon(icono, color=color_icono, size=24),
                        ft.Column([
                            ft.Text(desc, weight=ft.FontWeight.W_500, size=14, color=colores["texto"]),
                            ft.Text(f"{cat} · {fecha}", size=11, color=colores["texto_secundario"]),
                        ], expand=True, spacing=1),
                        ft.Column([
                            ft.Text(f"${monto:,.0f}", weight=ft.FontWeight.BOLD, color=color_icono, size=14),
                            ft.Row([
                                ft.IconButton(
                                    icon="edit_outlined",
                                    icon_color=colores["azul"],
                                    icon_size=16,
                                    tooltip="Editar",
                                    on_click=lambda e, m=mov: abrir_editar_movimiento(m)
                                ),
                                ft.IconButton(
                                    icon="delete_outline", 
                                    icon_color=colores["rojo"],
                                    icon_size=16,
                                    tooltip="Borrar",
                                    on_click=lambda e, x=id_mov, d=desc: confirmar_borrado("movimiento", x, d)
                                )
                            ], spacing=0)
                        ], alignment=ft.MainAxisAlignment.END, horizontal_alignment=ft.CrossAxisAlignment.END)
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    padding=10,
                    border_radius=10,
                    bgcolor=colores["tarjeta"],
                    border=ft.border.all(1, colores["borde"]),
                )
                lista_movimientos.controls.append(item)
        
        # Barra de búsqueda compacta
        barra_busqueda = ft.Container(
            content=ft.Row([
                ft.Container(content=input_busqueda, expand=True),
                ft.IconButton(
                    icon="filter_list",
                    icon_color=colores["azul"],
                    tooltip="Filtros",
                    on_click=lambda e: toggle_filtros()
                )
            ]),
            padding=ft.padding.only(left=10, right=10),
        )
        
        return ft.Column([
            crear_resumen_rapido(),
            crear_barra_progreso_mes(),
            crear_grafico_categorias_mini(),
            ft.Container(
                content=ft.Row([
                    ft.Text("📋 Últimos Movimientos", size=14, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                    ft.Text(f"({len(movimientos)})", size=12, color=colores["texto_secundario"])
                ]),
                padding=ft.padding.only(left=15, top=5, bottom=5)
            ),
            barra_busqueda,
            lista_movimientos
        ], spacing=0, expand=True, scroll=ft.ScrollMode.AUTO)
    
    def crear_vista_suscripciones():
        """Crea la vista de suscripciones"""
        colores = obtener_colores()
        lista_subs = ft.ListView(spacing=10, padding=10, expand=True)
        
        suscripciones = db.obtener_suscripciones()
        total = db.obtener_total_suscripciones()
        
        # Header con total
        header = ft.Container(
            content=ft.Column([
                ft.Text("📆 Suscripciones Activas", size=20, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                ft.Divider(height=10, color="transparent"),
                ft.Row([
                    ft.Text("Total mensual:", size=16, color=colores["texto_secundario"]),
                    ft.Text(f"${total:,.0f}", size=24, weight=ft.FontWeight.BOLD, color=colores["naranja"])
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
            ]),
            padding=20,
            bgcolor=colores["naranja_bg"],
            border_radius=15,
            margin=10,
            border=ft.border.all(1, colores["borde"])
        )
        
        if not suscripciones:
            lista_subs.controls.append(
                ft.Container(
                    content=ft.Text("No tienes suscripciones activas.\n¡Agrega tus servicios recurrentes!", 
                                   italic=True, text_align=ft.TextAlign.CENTER, size=14, color=colores["texto_secundario"]),
                    padding=40
                )
            )
        else:
            for sub in suscripciones:
                id_sub, nombre, monto, dia_cobro, activa = sub
                
                item = ft.Container(
                    content=ft.Row([
                        ft.Icon("subscriptions", color=colores["naranja"], size=28),
                        ft.Column([
                            ft.Text(nombre, weight=ft.FontWeight.BOLD, size=15, color=colores["texto"]),
                            ft.Text(f"Se cobra el día {dia_cobro} de cada mes", size=12, color=colores["texto_secundario"]),
                        ], expand=True, spacing=2),
                        ft.Column([
                            ft.Text(f"${monto:,.0f}/mes", weight=ft.FontWeight.BOLD, color=colores["naranja"], size=16),
                            ft.Row([
                                ft.IconButton(
                                    icon="edit_outlined",
                                    icon_color=colores["azul"],
                                    icon_size=18,
                                    tooltip="Editar",
                                    on_click=lambda e, s=sub: abrir_editar_suscripcion(s)
                                ),
                                ft.IconButton(
                                    icon="delete_outline", 
                                    icon_color=colores["rojo"],
                                    icon_size=18,
                                    tooltip="Eliminar",
                                    on_click=lambda e, x=id_sub, n=nombre: confirmar_borrado("suscripcion", x, n)
                                )
                            ], spacing=0)
                        ], alignment=ft.MainAxisAlignment.END, horizontal_alignment=ft.CrossAxisAlignment.END)
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    padding=12,
                    border_radius=12,
                    bgcolor=colores["tarjeta"],
                    border=ft.border.all(1, colores["borde"]),
                )
                lista_subs.controls.append(item)
        
        return ft.Column([header, lista_subs], spacing=0, expand=True)
    
    def crear_vista_prestamos():
        """Crea la vista de préstamos bancarios"""
        colores = obtener_colores()
        lista_prestamos = ft.ListView(spacing=10, padding=10, expand=True)
        
        prestamos = db.obtener_prestamos()
        total_cuotas = db.obtener_total_cuotas_prestamos()
        deuda_total = db.obtener_deuda_total()
        
        # Header con totales
        header = ft.Container(
            content=ft.Column([
                ft.Text("🏦 Préstamos Bancarios", size=20, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                ft.Divider(height=10, color="transparent"),
                ft.Row([
                    ft.Container(
                        content=ft.Column([
                            ft.Text("Cuotas/mes", size=12, color=colores["texto_secundario"]),
                            ft.Text(f"${total_cuotas:,.0f}", size=20, weight=ft.FontWeight.BOLD, color=colores["purple"])
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                        expand=True
                    ),
                    ft.Container(
                        content=ft.Column([
                            ft.Text("Deuda total", size=12, color=colores["texto_secundario"]),
                            ft.Text(f"${deuda_total:,.0f}", size=20, weight=ft.FontWeight.BOLD, color=colores["rojo"])
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                        expand=True
                    ),
                ], alignment=ft.MainAxisAlignment.SPACE_AROUND)
            ]),
            padding=20,
            bgcolor=colores["purple_bg"],
            border_radius=15,
            margin=10,
            border=ft.border.all(1, colores["borde"])
        )
        
        if not prestamos:
            lista_prestamos.controls.append(
                ft.Container(
                    content=ft.Text("No tienes préstamos registrados.\n¡Mantén control de tus deudas bancarias!", 
                                   italic=True, text_align=ft.TextAlign.CENTER, size=14, color=colores["texto_secundario"]),
                    padding=40
                )
            )
        else:
            for prestamo in prestamos:
                id_pres, banco, monto_total, monto_pagado, cuota_mensual, dia_pago, fecha_inicio, activo = prestamo
                
                saldo_pendiente = monto_total - monto_pagado
                porcentaje_pagado = (monto_pagado / monto_total * 100) if monto_total > 0 else 0
                
                item = ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon("account_balance", color=colores["purple"], size=28),
                            ft.Column([
                                ft.Text(banco, weight=ft.FontWeight.BOLD, size=15, color=colores["texto"]),
                                ft.Text(f"Cuota: ${cuota_mensual:,.0f}/mes · Día {dia_pago}", size=12, color=colores["texto_secundario"]),
                            ], expand=True, spacing=2),
                            ft.IconButton(
                                icon="delete_outline", 
                                icon_color=colores["rojo"],
                                icon_size=20,
                                tooltip="Eliminar",
                                on_click=lambda e, x=id_pres, b=banco: confirmar_borrado("prestamo", x, b)
                            )
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                        ft.Divider(height=5, color="transparent"),
                        ft.Row([
                            ft.Column([
                                ft.Text("Pagado", size=11, color=colores["texto_secundario"]),
                                ft.Text(f"${monto_pagado:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["verde"]),
                            ]),
                            ft.Column([
                                ft.Text("Pendiente", size=11, color=colores["texto_secundario"]),
                                ft.Text(f"${saldo_pendiente:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["rojo"]),
                            ]),
                            ft.Column([
                                ft.Text("Total", size=11, color=colores["texto_secundario"]),
                                ft.Text(f"${monto_total:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                            ]),
                        ], alignment=ft.MainAxisAlignment.SPACE_AROUND),
                        ft.ProgressBar(value=porcentaje_pagado/100, color=colores["purple"], bgcolor=colores["borde"]),
                        ft.Text(f"{porcentaje_pagado:.1f}% pagado", size=11, color=colores["purple"], text_align=ft.TextAlign.CENTER),
                        ft.ElevatedButton(
                            "Registrar Pago",
                            icon="payment",
                            on_click=lambda e, x=id_pres: abrir_registrar_pago(x),
                            bgcolor=colores["purple"],
                            color="white",
                            width=float("inf")
                        )
                    ], spacing=8),
                    padding=12,
                    border_radius=12,
                    bgcolor=colores["tarjeta"],
                    border=ft.border.all(1, colores["borde"]),
                )
                lista_prestamos.controls.append(item)
        
        return ft.Column([header, lista_prestamos], spacing=0, expand=True)
    
    def crear_vista_creditos():
        """Crea la vista de compras a crédito"""
        colores = obtener_colores()
        lista_creditos = ft.ListView(spacing=10, padding=10, expand=True)
        
        creditos = db.obtener_creditos()
        total_cuotas = db.obtener_total_cuotas_creditos()
        deuda_total = db.obtener_deuda_total_creditos()
        
        # Header con totales
        header = ft.Container(
            content=ft.Column([
                ft.Text("💳 Compras a Crédito", size=20, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                ft.Divider(height=10, color="transparent"),
                ft.Row([
                    ft.Container(
                        content=ft.Column([
                            ft.Text("Cuotas/mes", size=12, color=colores["texto_secundario"]),
                            ft.Text(f"${total_cuotas:,.0f}", size=20, weight=ft.FontWeight.BOLD, color=colores["indigo"])
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                        expand=True
                    ),
                    ft.Container(
                        content=ft.Column([
                            ft.Text("Deuda total", size=12, color=colores["texto_secundario"]),
                            ft.Text(f"${deuda_total:,.0f}", size=20, weight=ft.FontWeight.BOLD, color=colores["rojo"])
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                        expand=True
                    ),
                ], alignment=ft.MainAxisAlignment.SPACE_AROUND)
            ]),
            padding=20,
            bgcolor=colores["indigo_bg"],
            border_radius=15,
            margin=10
        )
        
        if not creditos:
            lista_creditos.controls.append(
                ft.Container(
                    content=ft.Text("No tienes compras a crédito.\n¡Controla tus compras en meses sin intereses!", 
                                   italic=True, text_align=ft.TextAlign.CENTER, size=14, color=colores["texto_secundario"]),
                    padding=40
                )
            )
        else:
            for credito in creditos:
                # credito = (id, descripcion, banco, monto_total, meses_sin_intereses, cuota_mensual, meses_pagados, fecha_compra, tasa_interes, pagado)
                id_cred, descripcion, banco, monto_total, meses_totales, cuota_mensual, meses_pagados, fecha_compra, tasa_interes, pagado = credito
                
                meses_restantes = meses_totales - meses_pagados
                saldo_pendiente = meses_restantes * cuota_mensual
                porcentaje_pagado = (meses_pagados / meses_totales * 100) if meses_totales > 0 else 0
                
                # Determinar si tiene intereses
                tipo_credito = "Sin intereses" if tasa_interes == 0 else f"Interés: {tasa_interes:.1f}% mensual"
                
                item = ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon("credit_card", color=colores["indigo"], size=28),
                            ft.Column([
                                ft.Text(descripcion, weight=ft.FontWeight.BOLD, size=15, color=colores["texto"]),
                                ft.Text(f"{banco} · {fecha_compra}", size=12, color=colores["texto_secundario"]),
                                ft.Text(tipo_credito, size=11, color=colores["indigo"], italic=True),
                            ], expand=True, spacing=2),
                            ft.IconButton(
                                icon="delete_outline", 
                                icon_color=colores["texto_secundario"],
                                icon_size=20,
                                tooltip="Eliminar",
                                on_click=lambda e, x=id_cred: borrar_credito(x)
                            )
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                        ft.Divider(height=5, color="transparent"),
                        ft.Row([
                            ft.Column([
                                ft.Text("Cuota mensual", size=11, color=colores["texto_secundario"]),
                                ft.Text(f"${cuota_mensual:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["indigo"]),
                            ]),
                            ft.Column([
                                ft.Text("Meses", size=11, color=colores["texto_secundario"]),
                                ft.Text(f"{meses_pagados}/{meses_totales}", size=14, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                            ]),
                            ft.Column([
                                ft.Text("Pendiente", size=11, color=colores["texto_secundario"]),
                                ft.Text(f"${saldo_pendiente:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["rojo"]),
                            ]),
                        ], alignment=ft.MainAxisAlignment.SPACE_AROUND),
                        ft.ProgressBar(value=porcentaje_pagado/100, color=colores["indigo"], bgcolor=colores["indigo_bg"]),
                        ft.Text(f"{porcentaje_pagado:.1f}% pagado · Faltan {meses_restantes} meses", size=11, color=colores["indigo"], text_align=ft.TextAlign.CENTER),
                        ft.ElevatedButton(
                            "Pagar Mensualidad",
                            icon="payment",
                            on_click=lambda e, x=id_cred: registrar_pago_credito_directo(x),
                            bgcolor=colores["indigo"],
                            color="white",
                            width=float("inf")
                        )
                    ], spacing=8),
                    padding=12,
                    border_radius=12,
                    bgcolor=colores["tarjeta"],
                    border=ft.border.all(1, colores["borde"]),
                    shadow=ft.BoxShadow(spread_radius=0, blur_radius=4, color="black12", offset=ft.Offset(0, 2))
                )
                lista_creditos.controls.append(item)
        
        return ft.Column([header, lista_creditos], spacing=0, expand=True)
    
    def crear_vista_ahorros():
        """Crea la vista de ahorros"""
        colores = obtener_colores()
        lista_ahorros = ft.ListView(spacing=10, padding=10, expand=True)
        
        ahorros = db.obtener_ahorros()
        total_ahorrado = db.obtener_total_ahorros()
        
        # Header con total
        header = ft.Container(
            content=ft.Column([
                ft.Text("🎯 Mis Ahorros", size=20, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                ft.Divider(height=10, color="transparent"),
                ft.Row([
                    ft.Text("Total ahorrado:", size=16, color=colores["texto_secundario"]),
                    ft.Text(f"${total_ahorrado:,.0f}", size=24, weight=ft.FontWeight.BOLD, color=colores["teal"])
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
            ]),
            padding=20,
            bgcolor=colores["teal_bg"],
            border_radius=15,
            margin=10
        )
        
        if not ahorros:
            lista_ahorros.controls.append(
                ft.Container(
                    content=ft.Text("No tienes metas de ahorro activas.\n¡Empieza a ahorrar para tus objetivos!", 
                                   italic=True, text_align=ft.TextAlign.CENTER, size=14, color=colores["texto_secundario"]),
                    padding=40
                )
            )
        else:
            for ahorro in ahorros:
                # ahorro = (id, nombre, meta, monto_actual, fecha_inicio, completado)
                id_aho, nombre, meta, monto_actual, fecha_inicio, completado = ahorro
                
                falta = meta - monto_actual
                porcentaje = (monto_actual / meta * 100) if meta > 0 else 0
                
                item = ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon("savings", color=colores["teal"], size=28),
                            ft.Column([
                                ft.Text(nombre, weight=ft.FontWeight.BOLD, size=15, color=colores["texto"]),
                                ft.Text(f"Desde {fecha_inicio}", size=12, color=colores["texto_secundario"]),
                            ], expand=True, spacing=2),
                            ft.IconButton(
                                icon="delete_outline", 
                                icon_color=colores["texto_secundario"],
                                icon_size=20,
                                tooltip="Eliminar",
                                on_click=lambda e, x=id_aho: borrar_ahorro(x)
                            )
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                        ft.Divider(height=5, color="transparent"),
                        ft.Row([
                            ft.Column([
                                ft.Text("Ahorrado", size=11, color=colores["texto_secundario"]),
                                ft.Text(f"${monto_actual:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["teal"]),
                            ]),
                            ft.Column([
                                ft.Text("Falta", size=11, color=colores["texto_secundario"]),
                                ft.Text(f"${falta:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["naranja"]),
                            ]),
                            ft.Column([
                                ft.Text("Meta", size=11, color=colores["texto_secundario"]),
                                ft.Text(f"${meta:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                            ]),
                        ], alignment=ft.MainAxisAlignment.SPACE_AROUND),
                        ft.ProgressBar(value=porcentaje/100, color=colores["teal"], bgcolor=colores["teal_bg"]),
                        ft.Text(f"{porcentaje:.1f}% completado", size=11, color=colores["teal"], text_align=ft.TextAlign.CENTER),
                        ft.Row([
                            ft.ElevatedButton(
                                "Agregar",
                                icon="add",
                                on_click=lambda e, x=id_aho: abrir_agregar_monto(x),
                                bgcolor=colores["teal"],
                                color="white",
                                expand=True
                            ),
                            ft.ElevatedButton(
                                "Retirar",
                                icon="remove",
                                on_click=lambda e, x=id_aho: abrir_retirar_monto(x),
                                bgcolor=colores["rojo"],
                                color="white",
                                expand=True
                            ),
                        ], spacing=10)
                    ], spacing=8),
                    padding=12,
                    border_radius=12,
                    bgcolor=colores["tarjeta"],
                    border=ft.border.all(1, colores["borde"]),
                    shadow=ft.BoxShadow(spread_radius=0, blur_radius=4, color="black12", offset=ft.Offset(0, 2))
                )
                lista_ahorros.controls.append(item)
        
        return ft.Column([header, lista_ahorros], spacing=0, expand=True)
    
    def crear_vista_bancos():
        """Crea la vista de cuentas bancarias"""
        colores = obtener_colores()
        lista_bancos = ft.ListView(spacing=10, padding=10, expand=True)
        
        cuentas = db.obtener_cuentas_bancarias()
        total_saldo = db.obtener_saldo_total_bancos()
        
        # Header con total
        header = ft.Container(
            content=ft.Column([
                ft.Text("🏦 Mis Cuentas Bancarias", size=20, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                ft.Divider(height=10, color="transparent"),
                ft.Row([
                    ft.Text("Saldo total:", size=16, color=colores["texto_secundario"]),
                    ft.Text(f"${total_saldo:,.0f}", size=24, weight=ft.FontWeight.BOLD, color=colores["cyan"])
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
            ]),
            padding=20,
            bgcolor=colores["cyan_bg"],
            border_radius=15,
            margin=10
        )
        
        if not cuentas:
            lista_bancos.controls.append(
                ft.Container(
                    content=ft.Text("No tienes cuentas registradas.\n¡Agrega tus cuentas bancarias!", 
                                   italic=True, text_align=ft.TextAlign.CENTER, size=14, color=colores["texto_secundario"]),
                    padding=40
                )
            )
        else:
            for cuenta in cuentas:
                # cuenta = (id, nombre_banco, tipo_cuenta, saldo, limite_credito, fecha_creacion, activa)
                id_cuenta, nombre_banco, tipo_cuenta, saldo, limite_credito, fecha_creacion, activa = cuenta
                
                # Iconos y colores según tipo de cuenta
                iconos = {
                    "debito": ("payment", colores["azul"]),
                    "credito": ("credit_card", colores["naranja"]),
                    "ahorro": ("savings", colores["verde"]),
                    "inversion": ("trending_up", colores["purple"])
                }
                icono, color = iconos.get(tipo_cuenta, ("account_balance", colores["cyan"]))
                
                # Mostrar información adicional para tarjetas de crédito
                info_adicional = ""
                if tipo_cuenta == "credito" and limite_credito > 0:
                    disponible_credito = limite_credito - abs(saldo)
                    info_adicional = f"Disponible: ${disponible_credito:,.0f} de ${limite_credito:,.0f}"
                
                item = ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(icono, color=color, size=28),
                            ft.Column([
                                ft.Text(nombre_banco, weight=ft.FontWeight.BOLD, size=15, color=colores["texto"]),
                                ft.Text(f"{tipo_cuenta.capitalize()} · {fecha_creacion}", size=12, color=colores["texto_secundario"]),
                            ], expand=True, spacing=2),
                            ft.IconButton(
                                icon="delete_outline", 
                                icon_color=colores["texto_secundario"],
                                icon_size=20,
                                tooltip="Eliminar",
                                on_click=lambda e, x=id_cuenta: borrar_cuenta_bancaria(x)
                            )
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                        ft.Divider(height=5, color="transparent"),
                        ft.Container(
                            content=ft.Column([
                                ft.Text("Saldo actual", size=12, color=colores["texto_secundario"]),
                                ft.Text(f"${saldo:,.0f}", size=24, weight=ft.FontWeight.BOLD, color=colores["cyan"]),
                                ft.Text(info_adicional, size=11, color=colores["texto_secundario"]) if info_adicional else ft.Container(),
                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                            padding=10,
                            bgcolor=colores["cyan_bg"],
                            border_radius=8,
                        ),
                        ft.Divider(height=5, color="transparent"),
                        ft.Row([
                            ft.ElevatedButton(
                                "Depositar",
                                icon="add",
                                on_click=lambda e, x=id_cuenta: abrir_depositar_banco(x),
                                bgcolor=colores["verde"],
                                color="white",
                                expand=True
                            ),
                            ft.ElevatedButton(
                                "Retirar",
                                icon="remove",
                                on_click=lambda e, x=id_cuenta: abrir_retirar_banco(x),
                                bgcolor=colores["rojo"],
                                color="white",
                                expand=True
                            ),
                        ], spacing=10)
                    ], spacing=8),
                    padding=12,
                    border_radius=12,
                    bgcolor=colores["tarjeta"],
                    border=ft.border.all(1, colores["borde"]),
                    shadow=ft.BoxShadow(spread_radius=0, blur_radius=4, color="black12", offset=ft.Offset(0, 2))
                )
                lista_bancos.controls.append(item)
        
        return ft.Column([header, lista_bancos], spacing=0, expand=True)
    
    def exportar_movimientos_a_excel(mes, anio):
        """Exporta los movimientos mensuales a un archivo Excel"""
        # Verificar si openpyxl está disponible
        if not EXCEL_DISPONIBLE:
            return False, "Exportación Excel no disponible en este dispositivo"
        
        try:
            # Obtener datos
            movimientos = db.obtener_movimientos_mensuales(mes, anio)
            ingresos_mes, gastos_mes = db.obtener_balance_mensual(mes, anio)
            total_subs = db.obtener_total_suscripciones()
            total_cuotas = db.obtener_total_cuotas_prestamos()
            total_cuotas_creditos = db.obtener_total_cuotas_creditos()
            balance_mes = ingresos_mes - gastos_mes
            
            meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            mes_nombre = meses[mes - 1]
            
            # Crear libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = f"{mes_nombre} {anio}"
            
            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            title_font = Font(bold=True, size=14)
            
            # Título
            ws.merge_cells('A1:E1')
            ws['A1'] = f"Reporte de Movimientos - {mes_nombre} {anio}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Resumen
            ws['A3'] = "RESUMEN DEL MES"
            ws['A3'].font = Font(bold=True, size=12)
            
            ws['A4'] = "Total Ingresos:"
            ws['B4'] = f"${ingresos_mes:,.2f}"
            ws['B4'].font = Font(color="008000", bold=True)
            
            ws['A5'] = "Total Gastos:"
            ws['B5'] = f"${gastos_mes:,.2f}"
            ws['B5'].font = Font(color="FF0000", bold=True)
            
            ws['A6'] = "Suscripciones:"
            ws['B6'] = f"${total_subs:,.2f}"
            
            ws['A7'] = "Cuotas Préstamos:"
            ws['B7'] = f"${total_cuotas:,.2f}"
            
            ws['A8'] = "Cuotas Créditos:"
            ws['B8'] = f"${total_cuotas_creditos:,.2f}"
            
            ws['A9'] = "Balance Final:"
            ws['B9'] = f"${balance_mes:,.2f}"
            ws['B9'].font = Font(bold=True, size=12)
            
            # Encabezados de movimientos
            ws['A11'] = "Fecha"
            ws['B11'] = "Tipo"
            ws['C11'] = "Categoría"
            ws['D11'] = "Descripción"
            ws['E11'] = "Monto"
            
            for cell in ['A11', 'B11', 'C11', 'D11', 'E11']:
                ws[cell].fill = header_fill
                ws[cell].font = header_font
                ws[cell].alignment = Alignment(horizontal='center')
            
            # Datos de movimientos
            fila = 12
            for mov in movimientos:
                id_mov, tipo, categoria, monto, descripcion, fecha = mov
                
                ws[f'A{fila}'] = fecha
                ws[f'B{fila}'] = tipo.upper()
                ws[f'C{fila}'] = categoria
                ws[f'D{fila}'] = descripcion
                ws[f'E{fila}'] = f"${monto:,.2f}"
                
                # Colorear según tipo
                if tipo == 'ingreso':
                    ws[f'B{fila}'].font = Font(color="008000")
                    ws[f'E{fila}'].font = Font(color="008000")
                else:
                    ws[f'B{fila}'].font = Font(color="FF0000")
                    ws[f'E{fila}'].font = Font(color="FF0000")
                
                fila += 1
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 15
            
            # Guardar archivo
            nombre_archivo = f"Movimientos_{mes_nombre}_{anio}.xlsx"
            ruta_documentos = os.path.join(os.path.expanduser('~'), 'Documents')
            ruta_completa = os.path.join(ruta_documentos, nombre_archivo)
            
            wb.save(ruta_completa)
            return True, ruta_completa
            
        except Exception as e:
            print(f"Error al exportar a Excel: {e}")
            return False, str(e)
    
    def crear_vista_balance_mensual():
        """Crea la vista de balance mensual con gráficos"""
        colores = obtener_colores()
        ahora = datetime.datetime.now()
        mes_actual = ahora.month
        anio_actual = ahora.year
        
        ingresos_mes, gastos_mes = db.obtener_balance_mensual(mes_actual, anio_actual)
        total_subs = db.obtener_total_suscripciones()
        total_cuotas = db.obtener_total_cuotas_prestamos()
        total_creditos = db.obtener_total_cuotas_creditos()
        balance_mes = ingresos_mes - gastos_mes - total_subs - total_cuotas - total_creditos
        
        meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
        mes_nombre = meses[mes_actual - 1]
        
        # Obtener gastos por categoría para el gráfico
        gastos_categoria = db.obtener_gastos_por_categoria(mes_actual, anio_actual)
        
        # Colores para las categorías
        colores_cat = {
            "Comida": "#FF6384", "Transporte": "#36A2EB", "Servicios": "#FFCE56",
            "Ocio": "#4BC0C0", "Salud": "#9966FF", "Salario": "#2ECC71",
            "Compras": "#FF9F40", "Educación": "#E74C3C", "Otro": "#95A5A6"
        }
        
        def exportar_excel(e):
            exito, mensaje = exportar_movimientos_a_excel(mes_actual, anio_actual)
            if exito:
                page.show_snack_bar(
                    ft.SnackBar(
                        content=ft.Text(f"✅ Excel exportado: {mensaje}"),
                        bgcolor=colores["verde"],
                        duration=5000
                    )
                )
            else:
                page.show_snack_bar(
                    ft.SnackBar(
                        content=ft.Text(f"❌ Error: {mensaje}"),
                        bgcolor=colores["rojo"],
                        duration=5000
                    )
                )
        
        # Crear gráfico de barras por categoría
        def crear_grafico_categorias():
            if not gastos_categoria:
                return ft.Container(
                    content=ft.Text("Sin gastos este mes", color=colores["texto_secundario"], italic=True),
                    padding=20,
                    alignment=ft.alignment.center
                )
            
            total_gastos = sum([g[1] for g in gastos_categoria])
            max_gasto = max([g[1] for g in gastos_categoria]) if gastos_categoria else 1
            
            barras = []
            for cat, monto in gastos_categoria:
                porcentaje = (monto / total_gastos * 100) if total_gastos > 0 else 0
                ancho_barra = (monto / max_gasto) if max_gasto > 0 else 0
                color = colores_cat.get(cat, "#95A5A6")
                
                barras.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Text(cat, size=12, weight=ft.FontWeight.W_500, expand=True, color=colores["texto"]),
                                ft.Text(f"${monto:,.0f}", size=12, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                                ft.Text(f"({porcentaje:.1f}%)", size=11, color=colores["texto_secundario"]),
                            ]),
                            ft.Container(
                                content=ft.Container(
                                    width=ancho_barra * 280,
                                    height=20,
                                    bgcolor=color,
                                    border_radius=5,
                                ),
                                bgcolor=colores["borde"],
                                border_radius=5,
                                width=280,
                                height=20,
                            )
                        ], spacing=5),
                        padding=ft.padding.only(bottom=10)
                    )
                )
            
            return ft.Container(
                content=ft.Column([
                    ft.Text("📊 Gastos por Categoría", size=16, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                    ft.Divider(height=10, color="transparent"),
                    *barras
                ]),
                padding=15,
                bgcolor=colores["gris_bg"],
                border_radius=12,
                margin=ft.margin.only(top=10, bottom=10)
            )
        
        # Crear mini gráfico de tendencia
        def crear_tendencia_mensual():
            datos_meses = db.obtener_balance_ultimos_meses(6)
            
            if not datos_meses:
                return ft.Container()
            
            max_valor = max([max(d["ingresos"], d["gastos"]) for d in datos_meses]) if datos_meses else 1
            
            barras_tendencia = []
            for d in datos_meses:
                altura_ing = (d["ingresos"] / max_valor * 60) if max_valor > 0 else 0
                altura_gas = (d["gastos"] / max_valor * 60) if max_valor > 0 else 0
                
                barras_tendencia.append(
                    ft.Column([
                        ft.Row([
                            ft.Container(width=15, height=altura_ing, bgcolor=colores["verde"], border_radius=3),
                            ft.Container(width=15, height=altura_gas, bgcolor=colores["rojo"], border_radius=3),
                        ], spacing=2, alignment=ft.MainAxisAlignment.CENTER),
                        ft.Text(d["mes"], size=10, color=colores["texto_secundario"])
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=5)
                )
            
            return ft.Container(
                content=ft.Column([
                    ft.Text("📈 Tendencia 6 Meses", size=16, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                    ft.Row([
                        ft.Row([ft.Container(width=10, height=10, bgcolor=colores["verde"], border_radius=2), ft.Text("Ingresos", size=10, color=colores["texto"])]),
                        ft.Row([ft.Container(width=10, height=10, bgcolor=colores["rojo"], border_radius=2), ft.Text("Gastos", size=10, color=colores["texto"])]),
                    ], spacing=20),
                    ft.Divider(height=10, color="transparent"),
                    ft.Row(barras_tendencia, alignment=ft.MainAxisAlignment.SPACE_AROUND),
                ]),
                padding=15,
                bgcolor=colores["azul_bg"],
                border_radius=12,
                margin=ft.margin.only(top=10)
            )
        
        return ft.Column([
            ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Text(f"📊 Balance de {mes_nombre} {anio_actual}", size=20, weight=ft.FontWeight.BOLD, expand=True, color=colores["texto"]),
                        ft.IconButton(
                            icon="download",
                            icon_color=colores["verde"],
                            tooltip="Exportar a Excel",
                            on_click=exportar_excel,
                            icon_size=28
                        )
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    ft.Divider(height=10, color="transparent"),
                    # Resumen en cards
                    ft.Row([
                        ft.Container(
                            content=ft.Column([
                                ft.Icon("trending_up", color=colores["verde"], size=24),
                                ft.Text("Ingresos", size=11, color=colores["texto_secundario"]),
                                ft.Text(f"${ingresos_mes:,.0f}", size=16, weight=ft.FontWeight.BOLD, color=colores["verde"]),
                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                            padding=12,
                            bgcolor=colores["verde_bg"],
                            border_radius=10,
                            expand=True
                        ),
                        ft.Container(
                            content=ft.Column([
                                ft.Icon("trending_down", color=colores["rojo"], size=24),
                                ft.Text("Gastos", size=11, color=colores["texto_secundario"]),
                                ft.Text(f"${gastos_mes:,.0f}", size=16, weight=ft.FontWeight.BOLD, color=colores["rojo"]),
                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                            padding=12,
                            bgcolor=colores["rojo_bg"],
                            border_radius=10,
                            expand=True
                        ),
                    ], spacing=10),
                    ft.Divider(height=10, color="transparent"),
                    ft.Row([
                        ft.Container(
                            content=ft.Column([
                                ft.Text("Suscripciones", size=10, color=colores["texto_secundario"]),
                                ft.Text(f"${total_subs:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["naranja"]),
                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                            padding=10,
                            bgcolor=colores["naranja_bg"],
                            border_radius=8,
                            expand=True
                        ),
                        ft.Container(
                            content=ft.Column([
                                ft.Text("Préstamos", size=10, color=colores["texto_secundario"]),
                                ft.Text(f"${total_cuotas:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["purple"]),
                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                            padding=10,
                            bgcolor=colores["purple_bg"],
                            border_radius=8,
                            expand=True
                        ),
                        ft.Container(
                            content=ft.Column([
                                ft.Text("Créditos", size=10, color=colores["texto_secundario"]),
                                ft.Text(f"${total_creditos:,.0f}", size=14, weight=ft.FontWeight.BOLD, color=colores["indigo"]),
                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                            padding=10,
                            bgcolor=colores["indigo_bg"],
                            border_radius=8,
                            expand=True
                        ),
                    ], spacing=8),
                    ft.Divider(height=15, color="transparent"),
                    # Balance final
                    ft.Container(
                        content=ft.Column([
                            ft.Text("💰 Balance Final del Mes", size=14, color=colores["texto_secundario"]),
                            ft.Text(f"${balance_mes:,.0f}", size=32, weight=ft.FontWeight.BOLD, 
                                   color=colores["verde"] if balance_mes >= 0 else colores["rojo"]),
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                        padding=20,
                        bgcolor=colores["tarjeta"],
                        border_radius=15,
                        border=ft.border.all(2, colores["verde"] if balance_mes >= 0 else colores["rojo"]),
                        shadow=ft.BoxShadow(spread_radius=1, blur_radius=8, color="black12", offset=ft.Offset(0, 2))
                    ),
                    # Gráfico de categorías
                    crear_grafico_categorias(),
                    # Tendencia mensual
                    crear_tendencia_mensual(),
                ]),
                padding=15,
                expand=True
            )
        ], spacing=0, expand=True, scroll=ft.ScrollMode.AUTO)
    
    def crear_resumen_balance():
        """Crea el widget de resumen de balance"""
        colores = obtener_colores()
        return ft.Container(
            content=ft.Column([
                ft.Text("Balance Total", size=14, color=colores["texto_secundario"], weight=ft.FontWeight.W_500),
                txt_balance_total,
                ft.Divider(height=5, color="transparent"),
                ft.Row([
                    ft.Container(
                        content=ft.Column([
                            ft.Icon("arrow_upward", color=colores["verde"], size=18),
                            ft.Text("Ingresos", size=11, color=colores["texto_secundario"]),
                            txt_ingresos
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                        padding=8,
                        bgcolor=colores["verde_bg"],
                        border_radius=8,
                        expand=True
                    ),
                    ft.Container(
                        content=ft.Column([
                            ft.Icon("arrow_downward", color=colores["rojo"], size=18),
                            ft.Text("Gastos", size=11, color=colores["texto_secundario"]),
                            txt_gastos
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                        padding=8,
                        bgcolor=colores["rojo_bg"],
                        border_radius=8,
                        expand=True
                    ),
                ], spacing=8),
                ft.Divider(height=5, color="transparent"),
                ft.Row([
                    ft.Container(
                        content=ft.Column([
                            ft.Icon("subscriptions", color=colores["naranja"], size=18),
                            ft.Text("Suscripciones", size=11, color=colores["texto_secundario"]),
                            txt_suscripciones
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                        padding=8,
                        bgcolor=colores["naranja_bg"],
                        border_radius=8,
                        expand=True
                    ),
                    ft.Container(
                        content=ft.Column([
                            ft.Icon("account_balance", color=colores["purple"], size=18),
                            ft.Text("Préstamos", size=11, color=colores["texto_secundario"]),
                            txt_prestamos
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                        padding=8,
                        bgcolor=colores["purple_bg"],
                        border_radius=8,
                        expand=True
                    ),
                ], spacing=8),
                ft.Divider(height=5, color="transparent"),
                ft.Row([
                    ft.Container(
                        content=ft.Column([
                            ft.Icon("credit_card", color=colores["indigo"], size=18),
                            ft.Text("Créditos", size=11, color=colores["texto_secundario"]),
                            txt_creditos
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                        padding=8,
                        bgcolor=colores["indigo_bg"],
                        border_radius=8,
                        expand=True
                    ),
                    ft.Container(
                        content=ft.Column([
                            ft.Icon("savings", color=colores["teal"], size=18),
                            ft.Text("Ahorros", size=11, color=colores["texto_secundario"]),
                            txt_ahorros
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                        padding=8,
                        bgcolor=colores["teal_bg"],
                        border_radius=8,
                        expand=True
                    ),
                ], spacing=8),
                ft.Divider(height=5, color="transparent"),
                ft.Row([
                    ft.Container(
                        content=ft.Column([
                            ft.Icon("account_balance", color=colores["cyan"], size=18),
                            ft.Text("En Bancos", size=11, color=colores["texto_secundario"]),
                            txt_bancos
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                        padding=8,
                        bgcolor=colores["cyan_bg"],
                        border_radius=8,
                        expand=True
                    ),
                    ft.Container(
                        content=ft.Column([
                            ft.Icon("account_balance_wallet", color=colores["azul"], size=18),
                            ft.Text("Disponible", size=11, color=colores["texto_secundario"]),
                            txt_disponible
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                        padding=8,
                        bgcolor=colores["azul_bg"],
                        border_radius=8,
                        expand=True
                    ),
                ], spacing=8),
            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=5),
            padding=15,
            margin=10,
            bgcolor=colores["tarjeta"],
            border_radius=15,
            shadow=ft.BoxShadow(spread_radius=0, blur_radius=8, color="black12", offset=ft.Offset(0, 2))
        )

    def guardar_movimiento(e):
        # Limpiar errores previos
        input_desc.error_text = None
        input_monto.error_text = None
        
        # Validar campos vacíos
        if not input_desc.value or not input_monto.value:
            input_desc.error_text = "Requerido" if not input_desc.value else None
            input_monto.error_text = "Requerido" if not input_monto.value else None
            page.update()
            return
        
        # Validar formato de número
        try:
            monto = float(input_monto.value)
        except ValueError:
            input_monto.error_text = "Debe ser un número válido"
            page.update()
            return
        
        # Validar monto positivo
        if monto <= 0:
            input_monto.error_text = "Debe ser mayor a 0"
            page.update()
            return

        # Intentar guardar en BD
        if db.agregar_movimiento(
            dropdown_tipo.value,
            dropdown_cat.value,
            monto,
            input_desc.value
        ):
            # Si usa banco, agregar o retirar el monto de la cuenta
            if dropdown_destino_movimiento.value == "banco" and dropdown_banco_movimiento.value:
                # Extraer el ID del banco del valor seleccionado
                id_cuenta = int(dropdown_banco_movimiento.value.split("_")[1])
                
                if dropdown_tipo.value == "ingreso":
                    # Agregar monto al banco
                    db.agregar_monto_cuenta(id_cuenta, monto)
                else:
                    # Retirar monto del banco (gasto)
                    db.retirar_monto_cuenta(id_cuenta, monto)
            
            # Limpiar campos y cerrar diálogo
            input_desc.value = ""
            input_monto.value = ""
            input_desc.error_text = None
            input_monto.error_text = None
            dropdown_banco_movimiento.visible = False
            bottom_sheet_movimiento.open = False
            actualizar_vista()
        else:
            # Mostrar error si falla el guardado
            input_desc.error_text = "Error al guardar. Intenta de nuevo."
            page.update()
    
    def guardar_suscripcion(e):
        input_sub_nombre.error_text = None
        input_sub_monto.error_text = None
        input_sub_dia.error_text = None
        
        if not input_sub_nombre.value or not input_sub_monto.value or not input_sub_dia.value:
            input_sub_nombre.error_text = "Requerido" if not input_sub_nombre.value else None
            input_sub_monto.error_text = "Requerido" if not input_sub_monto.value else None
            input_sub_dia.error_text = "Requerido" if not input_sub_dia.value else None
            page.update()
            return
        
        try:
            monto = float(input_sub_monto.value)
            dia = int(input_sub_dia.value)
        except ValueError:
            input_sub_monto.error_text = "Debe ser un número"
            page.update()
            return
        
        if monto <= 0:
            input_sub_monto.error_text = "Debe ser mayor a 0"
            page.update()
            return
        
        if dia < 1 or dia > 31:
            input_sub_dia.error_text = "Debe estar entre 1 y 31"
            page.update()
            return
        
        if db.agregar_suscripcion(input_sub_nombre.value, monto, dia):
            input_sub_nombre.value = ""
            input_sub_monto.value = ""
            input_sub_dia.value = ""
            bottom_sheet_suscripcion.open = False
            actualizar_vista()
        else:
            input_sub_nombre.error_text = "Error al guardar"
            page.update()
    
    def guardar_prestamo(e):
        input_prest_banco.error_text = None
        input_prest_monto_total.error_text = None
        input_prest_cuota.error_text = None
        input_prest_dia.error_text = None
        
        if not input_prest_banco.value or not input_prest_monto_total.value or not input_prest_cuota.value or not input_prest_dia.value:
            input_prest_banco.error_text = "Requerido" if not input_prest_banco.value else None
            input_prest_monto_total.error_text = "Requerido" if not input_prest_monto_total.value else None
            input_prest_cuota.error_text = "Requerido" if not input_prest_cuota.value else None
            input_prest_dia.error_text = "Requerido" if not input_prest_dia.value else None
            page.update()
            return
        
        try:
            monto_total = float(input_prest_monto_total.value)
            cuota = float(input_prest_cuota.value)
            dia = int(input_prest_dia.value)
        except ValueError:
            input_prest_monto_total.error_text = "Debe ser un número"
            page.update()
            return
        
        if monto_total <= 0 or cuota <= 0:
            input_prest_monto_total.error_text = "Debe ser mayor a 0"
            input_prest_cuota.error_text = "Debe ser mayor a 0"
            page.update()
            return
        
        if dia < 1 or dia > 31:
            input_prest_dia.error_text = "Debe estar entre 1 y 31"
            page.update()
            return
        
        if db.agregar_prestamo(input_prest_banco.value, monto_total, cuota, dia):
            input_prest_banco.value = ""
            input_prest_monto_total.value = ""
            input_prest_cuota.value = ""
            input_prest_dia.value = ""
            bottom_sheet_prestamo.open = False
            actualizar_vista()
        else:
            input_prest_banco.error_text = "Error al guardar"
            page.update()
    
    def guardar_credito(e):
        input_credito_desc.error_text = None
        input_credito_banco.error_text = None
        input_credito_monto.error_text = None
        input_credito_meses.error_text = None
        input_credito_interes.error_text = None
        
        if not input_credito_desc.value or not input_credito_banco.value or not input_credito_monto.value or not input_credito_meses.value:
            input_credito_desc.error_text = "Requerido" if not input_credito_desc.value else None
            input_credito_banco.error_text = "Requerido" if not input_credito_banco.value else None
            input_credito_monto.error_text = "Requerido" if not input_credito_monto.value else None
            input_credito_meses.error_text = "Requerido" if not input_credito_meses.value else None
            page.update()
            return
        
        try:
            monto = float(input_credito_monto.value)
            meses = int(input_credito_meses.value)
            tasa_interes = float(input_credito_interes.value) if input_credito_interes.value else 0
        except ValueError:
            input_credito_monto.error_text = "Debe ser un número"
            page.update()
            return
        
        if monto <= 0 or meses <= 0:
            input_credito_monto.error_text = "Debe ser mayor a 0" if monto <= 0 else None
            input_credito_meses.error_text = "Debe ser mayor a 0" if meses <= 0 else None
            page.update()
            return
        
        if tasa_interes < 0:
            input_credito_interes.error_text = "No puede ser negativo"
            page.update()
            return
        
        if db.agregar_credito(input_credito_desc.value, input_credito_banco.value, monto, meses, tasa_interes):
            input_credito_desc.value = ""
            input_credito_banco.value = ""
            input_credito_monto.value = ""
            input_credito_meses.value = ""
            input_credito_interes.value = "0"
            bottom_sheet_credito.open = False
            actualizar_vista()
        else:
            input_credito_desc.error_text = "Error al guardar"
            page.update()
    
    def guardar_ahorro(e):
        input_ahorro_nombre.error_text = None
        input_ahorro_meta.error_text = None
        
        if not input_ahorro_nombre.value or not input_ahorro_meta.value:
            input_ahorro_nombre.error_text = "Requerido" if not input_ahorro_nombre.value else None
            input_ahorro_meta.error_text = "Requerido" if not input_ahorro_meta.value else None
            page.update()
            return
        
        try:
            meta = float(input_ahorro_meta.value)
        except ValueError:
            input_ahorro_meta.error_text = "Debe ser un número"
            page.update()
            return
        
        if meta <= 0:
            input_ahorro_meta.error_text = "Debe ser mayor a 0"
            page.update()
            return
        
        if db.agregar_ahorro(input_ahorro_nombre.value, meta):
            input_ahorro_nombre.value = ""
            input_ahorro_meta.value = ""
            bottom_sheet_ahorro.open = False
            actualizar_vista()
        else:
            input_ahorro_nombre.error_text = "Error al guardar"
            page.update()
    
    def guardar_cuenta_bancaria(e):
        input_banco_nombre.error_text = None
        input_banco_saldo.error_text = None
        input_banco_limite.error_text = None
        
        if not input_banco_nombre.value:
            input_banco_nombre.error_text = "Requerido"
            page.update()
            return
        
        try:
            saldo = float(input_banco_saldo.value) if input_banco_saldo.value else 0
            limite = float(input_banco_limite.value) if input_banco_limite.value else 0
        except ValueError:
            input_banco_saldo.error_text = "Debe ser un número"
            page.update()
            return
        
        if db.agregar_cuenta_bancaria(input_banco_nombre.value, dropdown_tipo_cuenta.value, saldo, limite):
            input_banco_nombre.value = ""
            input_banco_saldo.value = "0"
            input_banco_limite.value = "0"
            dropdown_tipo_cuenta.value = "debito"
            bottom_sheet_banco.open = False
            actualizar_vista()
        else:
            input_banco_nombre.error_text = "Error al guardar"
            page.update()

    def borrar_movimiento(id_mov):
        if db.borrar_movimiento(id_mov):
            actualizar_vista()
    
    def borrar_suscripcion(id_sub):
        if db.borrar_suscripcion(id_sub):
            actualizar_vista()
    
    def borrar_prestamo(id_pres):
        if db.borrar_prestamo(id_pres):
            actualizar_vista()
    
    def borrar_ahorro(id_aho):
        if db.borrar_ahorro(id_aho):
            actualizar_vista()
    
    def borrar_credito(id_cred):
        if db.borrar_credito(id_cred):
            actualizar_vista()
    
    def registrar_pago_credito_directo(id_cred):
        if db.registrar_pago_credito(id_cred):
            actualizar_vista()
    
    def borrar_cuenta_bancaria(id_cuenta):
        if db.borrar_cuenta_bancaria(id_cuenta):
            actualizar_vista()

    # --- Elementos de Navegación y Estructura ---
    
    # Campos para suscripciones
    input_sub_nombre = ft.TextField(
        label="Nombre",
        hint_text="Ej: Netflix, Spotify...",
        color="black",
        text_size=16,
        border_color="orange700",
        focused_border_color="orange900"
    )
    input_sub_monto = ft.TextField(
        label="Monto mensual",
        keyboard_type=ft.KeyboardType.NUMBER,
        color="black",
        text_size=16,
        border_color="orange700",
        focused_border_color="orange900"
    )
    input_sub_dia = ft.TextField(
        label="Día de cobro (1-31)",
        keyboard_type=ft.KeyboardType.NUMBER,
        color="black",
        text_size=16,
        border_color="orange700",
        focused_border_color="orange900"
    )
    
    # Campos para préstamos
    input_prest_banco = ft.TextField(
        label="Banco",
        hint_text="Ej: Banco Nacional, BBVA...",
        color="black",
        text_size=16,
        border_color="purple700",
        focused_border_color="purple900"
    )
    input_prest_monto_total = ft.TextField(
        label="Monto total del préstamo",
        keyboard_type=ft.KeyboardType.NUMBER,
        color="black",
        text_size=16,
        border_color="purple700",
        focused_border_color="purple900"
    )
    input_prest_cuota = ft.TextField(
        label="Cuota mensual",
        keyboard_type=ft.KeyboardType.NUMBER,
        color="black",
        text_size=16,
        border_color="purple700",
        focused_border_color="purple900"
    )
    input_prest_dia = ft.TextField(
        label="Día de pago (1-31)",
        keyboard_type=ft.KeyboardType.NUMBER,
        color="black",
        text_size=16,
        border_color="purple700",
        focused_border_color="purple900"
    )
    
    # Campo para registrar pago de préstamo
    input_pago_monto = ft.TextField(
        label="Monto del pago",
        keyboard_type=ft.KeyboardType.NUMBER,
        color="black",
        text_size=16,
        border_color="purple700",
        focused_border_color="purple900"
    )
    
    # Campos para créditos
    input_credito_desc = ft.TextField(
        label="Descripción de la compra",
        hint_text="Ej: TV, Laptop, Refrigerador...",
        color="black",
        text_size=16,
        border_color="indigo700",
        focused_border_color="indigo900"
    )
    input_credito_banco = ft.TextField(
        label="Banco/Tarjeta",
        hint_text="Ej: BBVA, Santander, Liverpool...",
        color="black",
        text_size=16,
        border_color="indigo700",
        focused_border_color="indigo900"
    )
    input_credito_monto = ft.TextField(
        label="Monto total",
        keyboard_type=ft.KeyboardType.NUMBER,
        color="black",
        text_size=16,
        border_color="indigo700",
        focused_border_color="indigo900"
    )
    input_credito_meses = ft.TextField(
        label="Plazo en meses",
        keyboard_type=ft.KeyboardType.NUMBER,
        color="black",
        text_size=16,
        border_color="indigo700",
        focused_border_color="indigo900"
    )
    input_credito_interes = ft.TextField(
        label="Tasa de interés mensual % (0 = sin intereses)",
        keyboard_type=ft.KeyboardType.NUMBER,
        hint_text="Ej: 0, 2.5, 3.8",
        value="0",
        color="black",
        text_size=16,
        border_color="indigo700",
        focused_border_color="indigo900"
    )
    
    # Campos para ahorros
    input_ahorro_nombre = ft.TextField(
        label="Nombre del ahorro",
        hint_text="Ej: Vacaciones, Auto, Casa...",
        color="black",
        text_size=16,
        border_color="teal700",
        focused_border_color="teal900"
    )
    input_ahorro_meta = ft.TextField(
        label="Meta de ahorro",
        keyboard_type=ft.KeyboardType.NUMBER,
        color="black",
        text_size=16,
        border_color="teal700",
        focused_border_color="teal900"
    )
    
    # Campo para agregar/retirar monto de ahorro
    input_monto_ahorro = ft.TextField(
        label="Monto",
        keyboard_type=ft.KeyboardType.NUMBER,
        color="black",
        text_size=16,
        border_color="teal700",
        focused_border_color="teal900"
    )
    
    # Campos para cuentas bancarias
    input_banco_nombre = ft.TextField(
        label="Nombre del banco",
        hint_text="Ej: BBVA, Santander, Banorte...",
        color="black",
        text_size=16,
        border_color="cyan900",
        focused_border_color="cyan900"
    )
    dropdown_tipo_cuenta = ft.Dropdown(
        label="Tipo de cuenta",
        options=[
            ft.dropdown.Option("debito", "Débito"),
            ft.dropdown.Option("credito", "Crédito"),
            ft.dropdown.Option("ahorro", "Ahorro"),
            ft.dropdown.Option("inversion", "Inversión"),
        ],
        value="debito",
        color="black",
        border_color="cyan900"
    )
    input_banco_saldo = ft.TextField(
        label="Saldo inicial",
        keyboard_type=ft.KeyboardType.NUMBER,
        value="0",
        color="black",
        text_size=16,
        border_color="cyan900",
        focused_border_color="cyan900"
    )
    input_banco_limite = ft.TextField(
        label="Límite de crédito (solo para tarjetas)",
        keyboard_type=ft.KeyboardType.NUMBER,
        value="0",
        color="black",
        text_size=16,
        border_color="cyan900",
        focused_border_color="cyan900"
    )
    
    # Campo para depositar/retirar de cuenta bancaria
    input_monto_banco = ft.TextField(
        label="Monto",
        keyboard_type=ft.KeyboardType.NUMBER,
        color="black",
        text_size=16,
        border_color="cyan900",
        focused_border_color="cyan900"
    )

    # Dialogo para agregar movimiento
    bottom_sheet_movimiento = ft.BottomSheet(
        ft.Container(
            ft.Column(
                [
                    ft.Text("💸 Agregar Movimiento", size=20, weight=ft.FontWeight.BOLD),
                    dropdown_tipo,
                    dropdown_destino_movimiento,
                    dropdown_banco_movimiento,
                    dropdown_cat,
                    input_desc,
                    input_monto,
                    ft.ElevatedButton("Guardar", on_click=guardar_movimiento, width=float("inf"))
                ],
                tight=True,
                spacing=15,
                scroll=ft.ScrollMode.AUTO
            ),
            padding=20,
            border_radius=ft.border_radius.only(top_left=20, top_right=20)
        ),
        is_scroll_controlled=True,
        use_safe_area=True
    )
    
    # Dialogo para agregar suscripción
    bottom_sheet_suscripcion = ft.BottomSheet(
        ft.Container(
            ft.Column(
                [
                    ft.Text("📆 Agregar Suscripción", size=20, weight=ft.FontWeight.BOLD),
                    input_sub_nombre,
                    input_sub_monto,
                    input_sub_dia,
                    ft.ElevatedButton("Guardar", on_click=guardar_suscripcion, width=float("inf"))
                ],
                tight=True,
                spacing=15,
                scroll=ft.ScrollMode.AUTO
            ),
            padding=20,
            border_radius=ft.border_radius.only(top_left=20, top_right=20)
        ),
        is_scroll_controlled=True,
        use_safe_area=True
    )
    
    # Dialogo para agregar préstamo
    bottom_sheet_prestamo = ft.BottomSheet(
        ft.Container(
            ft.Column(
                [
                    ft.Text("🏦 Agregar Préstamo", size=20, weight=ft.FontWeight.BOLD),
                    input_prest_banco,
                    input_prest_monto_total,
                    input_prest_cuota,
                    input_prest_dia,
                    ft.ElevatedButton("Guardar", on_click=guardar_prestamo, width=float("inf"))
                ],
                tight=True,
                spacing=15,
                scroll=ft.ScrollMode.AUTO
            ),
            padding=20,
            border_radius=ft.border_radius.only(top_left=20, top_right=20)
        ),
        is_scroll_controlled=True,
        use_safe_area=True
    )
    
    # Dialogo para agregar compra a crédito
    bottom_sheet_credito = ft.BottomSheet(
        ft.Container(
            ft.Column(
                [
                    ft.Text("💳 Agregar Compra a Crédito", size=20, weight=ft.FontWeight.BOLD),
                    input_credito_desc,
                    input_credito_banco,
                    input_credito_monto,
                    input_credito_meses,
                    input_credito_interes,
                    ft.Text("ℹ️ Si es sin intereses, deja el campo en 0", size=11, color="grey600", italic=True),
                    ft.ElevatedButton("Guardar", on_click=guardar_credito, width=float("inf"))
                ],
                tight=True,
                spacing=12,
                scroll=ft.ScrollMode.AUTO
            ),
            padding=20,
            border_radius=ft.border_radius.only(top_left=20, top_right=20)
        ),
        is_scroll_controlled=True,
        use_safe_area=True
    )
    
    # Variable para almacenar el ID del préstamo al registrar pago
    prestamo_id_pago = [None]
    
    # Dialogo para registrar pago de préstamo
    bottom_sheet_pago_prestamo = ft.BottomSheet(
        ft.Container(
            ft.Column(
                [
                    ft.Text("💳 Registrar Pago", size=20, weight=ft.FontWeight.BOLD),
                    input_pago_monto,
                    ft.ElevatedButton("Confirmar Pago", on_click=lambda e: registrar_pago(e), width=float("inf"))
                ],
                tight=True,
                spacing=15,
                scroll=ft.ScrollMode.AUTO
            ),
            padding=20,
            border_radius=ft.border_radius.only(top_left=20, top_right=20)
        ),
        is_scroll_controlled=True,
        use_safe_area=True
    )
    
    def abrir_registrar_pago(id_prestamo):
        prestamo_id_pago[0] = id_prestamo
        input_pago_monto.value = ""
        input_pago_monto.error_text = None
        bottom_sheet_pago_prestamo.open = True
        page.update()
    
    def registrar_pago(e):
        input_pago_monto.error_text = None
        
        if not input_pago_monto.value:
            input_pago_monto.error_text = "Requerido"
            page.update()
            return
        
        try:
            monto = float(input_pago_monto.value)
        except ValueError:
            input_pago_monto.error_text = "Debe ser un número"
            page.update()
            return
        
        if monto <= 0:
            input_pago_monto.error_text = "Debe ser mayor a 0"
            page.update()
            return
        
        if db.registrar_pago_prestamo(prestamo_id_pago[0], monto):
            input_pago_monto.value = ""
            bottom_sheet_pago_prestamo.open = False
            actualizar_vista()
        else:
            input_pago_monto.error_text = "Error al registrar pago"
            page.update()
    
    # Dialogo para agregar ahorro
    bottom_sheet_ahorro = ft.BottomSheet(
        ft.Container(
            ft.Column(
                [
                    ft.Text("🎯 Crear Meta de Ahorro", size=20, weight=ft.FontWeight.BOLD),
                    input_ahorro_nombre,
                    input_ahorro_meta,
                    ft.ElevatedButton("Crear Meta", on_click=guardar_ahorro, width=float("inf"))
                ],
                tight=True,
                spacing=15,
                scroll=ft.ScrollMode.AUTO
            ),
            padding=20,
            border_radius=ft.border_radius.only(top_left=20, top_right=20)
        ),
        is_scroll_controlled=True,
        use_safe_area=True
    )
    
    # Variable para almacenar el ID del ahorro al agregar/retirar monto
    ahorro_id_operacion = [None]
    operacion_tipo = ["agregar"]  # "agregar" o "retirar"
    
    # Dialogo para agregar/retirar monto de ahorro
    bottom_sheet_monto_ahorro = ft.BottomSheet(
        ft.Container(
            ft.Column(
                [
                    ft.Text("💰 Modificar Ahorro", size=20, weight=ft.FontWeight.BOLD),
                    input_monto_ahorro,
                    ft.ElevatedButton("Confirmar", on_click=lambda e: confirmar_operacion_ahorro(e), width=float("inf"))
                ],
                tight=True,
                spacing=15,
                scroll=ft.ScrollMode.AUTO
            ),
            padding=20,
            border_radius=ft.border_radius.only(top_left=20, top_right=20)
        ),
        is_scroll_controlled=True,
        use_safe_area=True
    )
    
    def abrir_agregar_monto(id_ahorro):
        ahorro_id_operacion[0] = id_ahorro
        operacion_tipo[0] = "agregar"
        input_monto_ahorro.value = ""
        input_monto_ahorro.error_text = None
        input_monto_ahorro.label = "Monto a agregar"
        bottom_sheet_monto_ahorro.open = True
        page.update()
    
    def abrir_retirar_monto(id_ahorro):
        ahorro_id_operacion[0] = id_ahorro
        operacion_tipo[0] = "retirar"
        input_monto_ahorro.value = ""
        input_monto_ahorro.error_text = None
        input_monto_ahorro.label = "Monto a retirar"
        bottom_sheet_monto_ahorro.open = True
        page.update()
    
    def confirmar_operacion_ahorro(e):
        input_monto_ahorro.error_text = None
        
        if not input_monto_ahorro.value:
            input_monto_ahorro.error_text = "Requerido"
            page.update()
            return
        
        try:
            monto = float(input_monto_ahorro.value)
        except ValueError:
            input_monto_ahorro.error_text = "Debe ser un número"
            page.update()
            return
        
        if monto <= 0:
            input_monto_ahorro.error_text = "Debe ser mayor a 0"
            page.update()
            return
        
        exito = False
        if operacion_tipo[0] == "agregar":
            exito = db.agregar_monto_ahorro(ahorro_id_operacion[0], monto)
        else:
            exito = db.retirar_monto_ahorro(ahorro_id_operacion[0], monto)
        
        if exito:
            input_monto_ahorro.value = ""
            bottom_sheet_monto_ahorro.open = False
            actualizar_vista()
        else:
            input_monto_ahorro.error_text = "Error en la operación"
            page.update()
    
    # Dialogo para agregar cuenta bancaria
    bottom_sheet_banco = ft.BottomSheet(
        ft.Container(
            ft.Column(
                [
                    ft.Text("🏦 Agregar Cuenta Bancaria", size=20, weight=ft.FontWeight.BOLD),
                    input_banco_nombre,
                    dropdown_tipo_cuenta,
                    input_banco_saldo,
                    input_banco_limite,
                    ft.Text("ℹ️ El límite solo aplica para tarjetas de crédito", size=11, color="grey600", italic=True),
                    ft.ElevatedButton("Guardar", on_click=guardar_cuenta_bancaria, width=float("inf"))
                ],
                tight=True,
                spacing=12,
                scroll=ft.ScrollMode.AUTO
            ),
            padding=20,
            border_radius=ft.border_radius.only(top_left=20, top_right=20)
        ),
        is_scroll_controlled=True,
        use_safe_area=True
    )
    
    # Variable para almacenar el ID de la cuenta al depositar/retirar
    cuenta_id_operacion = [None]
    operacion_tipo_banco = ["depositar"]  # "depositar" o "retirar"
    
    # Dialogo para depositar/retirar de cuenta bancaria
    bottom_sheet_monto_banco = ft.BottomSheet(
        ft.Container(
            ft.Column(
                [
                    ft.Text("💰 Modificar Saldo", size=20, weight=ft.FontWeight.BOLD),
                    input_monto_banco,
                    ft.ElevatedButton("Confirmar", on_click=lambda e: confirmar_operacion_banco(e), width=float("inf"))
                ],
                tight=True,
                spacing=15,
                scroll=ft.ScrollMode.AUTO
            ),
            padding=20,
            border_radius=ft.border_radius.only(top_left=20, top_right=20)
        ),
        is_scroll_controlled=True,
        use_safe_area=True
    )
    
    def abrir_depositar_banco(id_cuenta):
        cuenta_id_operacion[0] = id_cuenta
        operacion_tipo_banco[0] = "depositar"
        input_monto_banco.value = ""
        input_monto_banco.error_text = None
        input_monto_banco.label = "Monto a depositar"
        bottom_sheet_monto_banco.open = True
        page.update()
    
    def abrir_retirar_banco(id_cuenta):
        cuenta_id_operacion[0] = id_cuenta
        operacion_tipo_banco[0] = "retirar"
        input_monto_banco.value = ""
        input_monto_banco.error_text = None
        input_monto_banco.label = "Monto a retirar"
        bottom_sheet_monto_banco.open = True
        page.update()
    
    def confirmar_operacion_banco(e):
        input_monto_banco.error_text = None
        
        if not input_monto_banco.value:
            input_monto_banco.error_text = "Requerido"
            page.update()
            return
        
        try:
            monto = float(input_monto_banco.value)
        except ValueError:
            input_monto_banco.error_text = "Debe ser un número"
            page.update()
            return
        
        if monto <= 0:
            input_monto_banco.error_text = "Debe ser mayor a 0"
            page.update()
            return
        
        exito = False
        if operacion_tipo_banco[0] == "depositar":
            exito = db.agregar_monto_cuenta(cuenta_id_operacion[0], monto)
        else:
            exito = db.retirar_monto_cuenta(cuenta_id_operacion[0], monto)
        
        if exito:
            input_monto_banco.value = ""
            bottom_sheet_monto_banco.open = False
            actualizar_vista()
        else:
            input_monto_banco.error_text = "Error en la operación"
            page.update()

    def abrir_agregar(e):
        nonlocal vista_actual
        
        if vista_actual == "suscripciones":
            # Limpiar campos de suscripción
            input_sub_nombre.value = ""
            input_sub_monto.value = ""
            input_sub_dia.value = ""
            input_sub_nombre.error_text = None
            input_sub_monto.error_text = None
            input_sub_dia.error_text = None
            bottom_sheet_suscripcion.open = True
        elif vista_actual == "prestamos":
            # Limpiar campos de préstamo
            input_prest_banco.value = ""
            input_prest_monto_total.value = ""
            input_prest_cuota.value = ""
            input_prest_dia.value = ""
            input_prest_banco.error_text = None
            input_prest_monto_total.error_text = None
            input_prest_cuota.error_text = None
            input_prest_dia.error_text = None
            bottom_sheet_prestamo.open = True
        elif vista_actual == "creditos":
            # Limpiar campos de crédito
            input_credito_desc.value = ""
            input_credito_banco.value = ""
            input_credito_monto.value = ""
            input_credito_meses.value = ""
            input_credito_interes.value = "0"
            input_credito_desc.error_text = None
            input_credito_banco.error_text = None
            input_credito_monto.error_text = None
            input_credito_meses.error_text = None
            input_credito_interes.error_text = None
            bottom_sheet_credito.open = True
        elif vista_actual == "ahorros":
            # Limpiar campos de ahorro
            input_ahorro_nombre.value = ""
            input_ahorro_meta.value = ""
            input_ahorro_nombre.error_text = None
            input_ahorro_meta.error_text = None
            bottom_sheet_ahorro.open = True
        elif vista_actual == "bancos":
            # Limpiar campos de cuenta bancaria
            input_banco_nombre.value = ""
            input_banco_saldo.value = "0"
            input_banco_limite.value = "0"
            dropdown_tipo_cuenta.value = "debito"
            input_banco_nombre.error_text = None
            input_banco_saldo.error_text = None
            input_banco_limite.error_text = None
            bottom_sheet_banco.open = True
        else:
            # Limpiar campos de movimiento
            input_desc.value = ""
            input_monto.value = ""
            input_desc.error_text = None
            input_monto.error_text = None
            dropdown_tipo.value = "gasto"
            dropdown_cat.value = "Comida"
            dropdown_destino_movimiento.value = "efectivo"
            dropdown_banco_movimiento.visible = False
            actualizar_opciones_destino()
            bottom_sheet_movimiento.open = True
        
        page.update()

    # =====================================================
    # FUNCIONES DE EDICIÓN
    # =====================================================
    
    def abrir_editar_movimiento(mov):
        """Abre el formulario para editar un movimiento"""
        if len(mov) == 7:
            id_mov, tipo, cat, monto, desc, fecha, modo = mov
        else:
            id_mov, tipo, cat, monto, desc, fecha = mov
        
        registro_editando[0] = "movimiento"
        registro_editando[1] = id_mov
        
        input_desc.value = desc
        input_monto.value = str(monto)
        dropdown_tipo.value = tipo
        dropdown_cat.value = cat
        
        bottom_sheet_movimiento.open = True
        page.update()
    
    def abrir_editar_suscripcion(sub):
        """Abre el formulario para editar una suscripción"""
        id_sub, nombre, monto, dia_cobro, activa = sub
        
        registro_editando[0] = "suscripcion"
        registro_editando[1] = id_sub
        
        input_sub_nombre.value = nombre
        input_sub_monto.value = str(monto)
        input_sub_dia.value = str(dia_cobro)
        
        bottom_sheet_suscripcion.open = True
        page.update()
    
    def abrir_editar_prestamo(pres):
        """Abre el formulario para editar un préstamo"""
        id_pres, banco, monto_total, monto_pagado, cuota_mensual, dia_pago, fecha_inicio, activo = pres
        
        registro_editando[0] = "prestamo"
        registro_editando[1] = id_pres
        
        input_prest_banco.value = banco
        input_prest_monto_total.value = str(monto_total)
        input_prest_cuota.value = str(cuota_mensual)
        input_prest_dia.value = str(dia_pago)
        
        bottom_sheet_prestamo.open = True
        page.update()
    
    def abrir_editar_ahorro(aho):
        """Abre el formulario para editar un ahorro"""
        id_aho, nombre, meta, monto_actual, fecha_inicio, completado = aho
        
        registro_editando[0] = "ahorro"
        registro_editando[1] = id_aho
        
        input_ahorro_nombre.value = nombre
        input_ahorro_meta.value = str(meta)
        
        bottom_sheet_ahorro.open = True
        page.update()
    
    # =====================================================
    # VISTA DE CONFIGURACIÓN
    # =====================================================
    
    def crear_vista_configuracion():
        """Crea la vista de configuración"""
        colores = obtener_colores()
        tema_actual = db.obtener_tema()
        es_oscuro = tema_actual == "dark"
        
        def cambiar_tema(e):
            nuevo_tema = "dark" if e.control.value else "light"
            db.guardar_tema(nuevo_tema)
            page.theme_mode = ft.ThemeMode.DARK if nuevo_tema == "dark" else ft.ThemeMode.LIGHT
            page.update()
            actualizar_vista()
        
        def cambiar_pin(e):
            """Abre el diálogo para cambiar PIN"""
            # Limpiar campos
            for p in pin_inputs:
                p.value = ""
            txt_pin_titulo.value = "Crear nuevo PIN"
            txt_pin_mensaje.value = ""
            contenedor_login.visible = True
            contenedor_app.visible = False
            btn_saltar_pin.visible = True
            page.update()
        
        def exportar_backup(e):
            """Exporta todos los datos"""
            try:
                datos = db.exportar_datos()
                if datos:
                    nombre_archivo = f"JFinanzas_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                    ruta_documentos = os.path.join(os.path.expanduser('~'), 'Documents')
                    ruta_completa = os.path.join(ruta_documentos, nombre_archivo)
                    
                    with open(ruta_completa, 'w', encoding='utf-8') as f:
                        f.write(datos)
                    
                    page.show_snack_bar(
                        ft.SnackBar(
                            content=ft.Text(f"✅ Backup guardado en: {ruta_completa}"),
                            bgcolor=colores["verde"],
                            duration=5000
                        )
                    )
                else:
                    page.show_snack_bar(
                        ft.SnackBar(content=ft.Text("❌ Error al exportar"), bgcolor=colores["rojo"])
                    )
            except Exception as ex:
                page.show_snack_bar(
                    ft.SnackBar(content=ft.Text(f"❌ Error: {ex}"), bgcolor=colores["rojo"])
                )
        
        # FilePicker para importar backup
        def resultado_file_picker(e: ft.FilePickerResultEvent):
            if e.files and len(e.files) > 0:
                archivo = e.files[0]
                try:
                    with open(archivo.path, 'r', encoding='utf-8') as f:
                        json_data = f.read()
                    
                    if db.importar_datos(json_data):
                        page.show_snack_bar(
                            ft.SnackBar(
                                content=ft.Text("✅ Datos importados correctamente. Reinicia la app para ver los cambios."),
                                bgcolor=colores["verde"],
                                duration=5000
                            )
                        )
                        actualizar_vista()
                    else:
                        page.show_snack_bar(
                            ft.SnackBar(content=ft.Text("❌ Error al importar datos"), bgcolor=colores["rojo"])
                        )
                except Exception as ex:
                    page.show_snack_bar(
                        ft.SnackBar(content=ft.Text(f"❌ Error: {ex}"), bgcolor=colores["rojo"])
                    )
        
        file_picker = ft.FilePicker(on_result=resultado_file_picker)
        page.overlay.append(file_picker)
        
        def importar_backup(e):
            """Abre el selector de archivos para importar"""
            file_picker.pick_files(
                allowed_extensions=["json"],
                dialog_title="Selecciona el archivo de backup",
                file_type=ft.FilePickerFileType.CUSTOM
            )
        
        return ft.Column([
            ft.Container(
                content=ft.Column([
                    ft.Text("⚙️ Configuración", size=24, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                    ft.Divider(height=20, color=colores["borde"]),
                    
                    # Tema
                    ft.Container(
                        content=ft.Row([
                            ft.Icon("dark_mode", color=colores["purple"]),
                            ft.Column([
                                ft.Text("Modo Oscuro", weight=ft.FontWeight.BOLD, color=colores["texto"]),
                                ft.Text("Cambia la apariencia de la app", size=12, color=colores["texto_secundario"]),
                            ], expand=True, spacing=2),
                            ft.Switch(value=es_oscuro, on_change=cambiar_tema)
                        ]),
                        padding=15,
                        bgcolor=colores["purple_bg"],
                        border_radius=10,
                    ),
                    ft.Divider(height=10, color="transparent"),
                    
                    # Seguridad
                    ft.Container(
                        content=ft.Row([
                            ft.Icon("lock", color=colores["azul"]),
                            ft.Column([
                                ft.Text("Cambiar PIN", weight=ft.FontWeight.BOLD, color=colores["texto"]),
                                ft.Text("Modifica tu PIN de acceso", size=12, color=colores["texto_secundario"]),
                            ], expand=True, spacing=2),
                            ft.IconButton(icon="chevron_right", on_click=cambiar_pin, icon_color=colores["texto"])
                        ]),
                        padding=15,
                        bgcolor=colores["azul_bg"],
                        border_radius=10,
                    ),
                    ft.Divider(height=10, color="transparent"),
                    
                    # Backup - Exportar
                    ft.Container(
                        content=ft.Row([
                            ft.Icon("backup", color=colores["verde"]),
                            ft.Column([
                                ft.Text("Exportar Backup", weight=ft.FontWeight.BOLD, color=colores["texto"]),
                                ft.Text("Guarda todos tus datos en JSON", size=12, color=colores["texto_secundario"]),
                            ], expand=True, spacing=2),
                            ft.IconButton(icon="download", on_click=exportar_backup, icon_color=colores["texto"])
                        ]),
                        padding=15,
                        bgcolor=colores["verde_bg"],
                        border_radius=10,
                    ),
                    ft.Divider(height=10, color="transparent"),
                    
                    # Backup - Importar
                    ft.Container(
                        content=ft.Row([
                            ft.Icon("cloud_upload", color=colores["cyan"]),
                            ft.Column([
                                ft.Text("Importar Backup", weight=ft.FontWeight.BOLD, color=colores["texto"]),
                                ft.Text("Restaura datos desde archivo JSON", size=12, color=colores["texto_secundario"]),
                            ], expand=True, spacing=2),
                            ft.IconButton(icon="upload_file", on_click=importar_backup, icon_color=colores["texto"])
                        ]),
                        padding=15,
                        bgcolor=colores["cyan_bg"],
                        border_radius=10,
                    ),
                    ft.Divider(height=20, color="transparent"),
                    
                    # Info de la app
                    ft.Container(
                        content=ft.Column([
                            ft.Text("📱 Mis Finanzas", size=18, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                            ft.Text("Versión 2.0", size=14, color=colores["texto_secundario"]),
                            ft.Text("Gestiona tus finanzas de forma inteligente", size=12, color=colores["texto_secundario"]),
                            ft.Divider(height=10, color="transparent"),
                            ft.Row([
                                ft.Icon("code", size=16, color=colores["texto_secundario"]),
                                ft.Text("Desarrollado con Flet + Python", size=11, color=colores["texto_secundario"]),
                            ], spacing=5)
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                        padding=20,
                        bgcolor=colores["gris_bg"],
                        border_radius=15,
                        alignment=ft.alignment.center
                    ),
                ]),
                padding=20,
                expand=True
            )
        ], spacing=0, expand=True, scroll=ft.ScrollMode.AUTO)
    
    # =====================================================
    # VISTA DE PRESUPUESTOS
    # =====================================================
    
    def crear_vista_presupuestos():
        """Crea la vista de presupuestos por categoría"""
        colores = obtener_colores()
        lista_presupuestos = ft.ListView(spacing=10, padding=10, expand=True)
        
        presupuestos = db.obtener_presupuestos()
        
        # Header
        header = ft.Container(
            content=ft.Column([
                ft.Text("📋 Presupuestos por Categoría", size=20, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                ft.Divider(height=10, color="transparent"),
                ft.Text("Define límites de gasto para cada categoría y controla mejor tus finanzas.", 
                       size=14, color=colores["texto_secundario"]),
            ]),
            padding=20,
            bgcolor=colores["naranja_bg"],
            border_radius=15,
            margin=10
        )
        
        categorias = ["Comida", "Transporte", "Servicios", "Ocio", "Salud", "Compras", "Educación", "Otro"]
        
        for cat in categorias:
            # Buscar si hay presupuesto definido
            presupuesto_cat = next((p for p in presupuestos if p[1] == cat), None)
            limite = presupuesto_cat[2] if presupuesto_cat else 0
            id_pres = presupuesto_cat[0] if presupuesto_cat else None
            
            gasto_actual = db.obtener_gasto_categoria_mes(cat)
            porcentaje = (gasto_actual / limite * 100) if limite > 0 else 0
            
            # Color según el porcentaje
            if porcentaje >= 100:
                color_progreso = colores["rojo"]
                estado = "⚠️ Excedido"
            elif porcentaje >= 80:
                color_progreso = colores["naranja"]
                estado = "⚡ Cerca del límite"
            else:
                color_progreso = colores["verde"]
                estado = "✅ Dentro del presupuesto"
            
            item = ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Text(cat, weight=ft.FontWeight.BOLD, size=15, color=colores["texto"]),
                        ft.Text(estado if limite > 0 else "Sin límite", size=12, 
                               color=color_progreso if limite > 0 else colores["texto_secundario"]),
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    ft.Row([
                        ft.Text(f"Gastado: ${gasto_actual:,.0f}", size=13, color=colores["texto"]),
                        ft.Text(f"Límite: ${limite:,.0f}" if limite > 0 else "No definido", 
                               size=13, color=colores["texto_secundario"]),
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    ft.ProgressBar(
                        value=min(porcentaje/100, 1) if limite > 0 else 0,
                        color=color_progreso,
                        bgcolor=colores["borde"]
                    ) if limite > 0 else ft.Container(),
                    ft.Row([
                        ft.TextField(
                            hint_text="Límite",
                            keyboard_type=ft.KeyboardType.NUMBER,
                            width=120,
                            height=40,
                            text_size=14,
                            data=cat
                        ),
                        ft.ElevatedButton(
                            "Guardar",
                            bgcolor=colores["naranja"],
                            color="white",
                            height=40,
                            on_click=lambda e, c=cat: guardar_presupuesto_cat(e, c)
                        ),
                        ft.IconButton(
                            icon="delete",
                            icon_color=colores["rojo"],
                            tooltip="Eliminar límite",
                            on_click=lambda e, i=id_pres, c=cat: confirmar_borrado("presupuesto", i, f"presupuesto de {c}") if i else None,
                            visible=limite > 0
                        )
                    ], spacing=10)
                ], spacing=8),
                padding=15,
                bgcolor=colores["tarjeta"],
                border_radius=12,
                border=ft.border.all(1, colores["borde"])
            )
            lista_presupuestos.controls.append(item)
        
        return ft.Column([header, lista_presupuestos], spacing=0, expand=True)
    
    def guardar_presupuesto_cat(e, categoria):
        """Guarda el presupuesto de una categoría"""
        # Buscar el TextField correspondiente
        for control in e.control.parent.controls:
            if isinstance(control, ft.TextField) and control.data == categoria:
                try:
                    limite = float(control.value)
                    if limite > 0:
                        db.agregar_presupuesto(categoria, limite)
                        actualizar_vista()
                        page.show_snack_bar(
                            ft.SnackBar(content=ft.Text(f"✅ Presupuesto de {categoria} guardado"), bgcolor="green")
                        )
                except:
                    page.show_snack_bar(
                        ft.SnackBar(content=ft.Text("❌ Ingresa un número válido"), bgcolor="red")
                    )
                break
    
    # =====================================================
    # VISTA DE TRANSFERENCIAS
    # =====================================================
    
    def crear_vista_transferencias():
        """Crea la vista de transferencias entre cuentas"""
        colores = obtener_colores()
        cuentas = db.obtener_cuentas_bancarias()
        transferencias = db.obtener_transferencias()
        
        # Header con formulario de transferencia
        opciones_cuentas = [ft.dropdown.Option(str(c[0]), f"{c[1]} ({c[2]})") for c in cuentas]
        
        dropdown_origen = ft.Dropdown(
            label="Cuenta origen",
            options=opciones_cuentas,
            width=150
        )
        
        dropdown_destino = ft.Dropdown(
            label="Cuenta destino",
            options=opciones_cuentas,
            width=150
        )
        
        input_monto_trans = ft.TextField(
            label="Monto",
            keyboard_type=ft.KeyboardType.NUMBER,
            width=120
        )
        
        def realizar_transferencia_click(e):
            if not dropdown_origen.value or not dropdown_destino.value or not input_monto_trans.value:
                page.show_snack_bar(
                    ft.SnackBar(content=ft.Text("❌ Completa todos los campos"), bgcolor=colores["rojo"])
                )
                return
            
            if dropdown_origen.value == dropdown_destino.value:
                page.show_snack_bar(
                    ft.SnackBar(content=ft.Text("❌ Las cuentas deben ser diferentes"), bgcolor=colores["rojo"])
                )
                return
            
            try:
                monto = float(input_monto_trans.value)
                if monto <= 0:
                    raise ValueError()
                
                if db.realizar_transferencia(int(dropdown_origen.value), int(dropdown_destino.value), monto):
                    page.show_snack_bar(
                        ft.SnackBar(content=ft.Text("✅ Transferencia realizada"), bgcolor=colores["verde"])
                    )
                    dropdown_origen.value = None
                    dropdown_destino.value = None
                    input_monto_trans.value = ""
                    actualizar_vista()
                else:
                    page.show_snack_bar(
                        ft.SnackBar(content=ft.Text("❌ Error en la transferencia"), bgcolor=colores["rojo"])
                    )
            except:
                page.show_snack_bar(
                    ft.SnackBar(content=ft.Text("❌ Monto inválido"), bgcolor=colores["rojo"])
                )
        
        header = ft.Container(
            content=ft.Column([
                ft.Text("🔄 Transferir entre Cuentas", size=20, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                ft.Divider(height=10, color="transparent"),
                ft.Row([dropdown_origen, ft.Icon("arrow_forward", color=colores["texto"]), dropdown_destino], 
                       alignment=ft.MainAxisAlignment.CENTER, spacing=10),
                ft.Row([
                    input_monto_trans,
                    ft.ElevatedButton("Transferir", icon="send", bgcolor=colores["cyan"], color="white",
                                     on_click=realizar_transferencia_click)
                ], alignment=ft.MainAxisAlignment.CENTER, spacing=10)
            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            padding=20,
            bgcolor=colores["cyan_bg"],
            border_radius=15,
            margin=10
        )
        
        # Historial de transferencias
        lista_trans = ft.ListView(spacing=10, padding=10, expand=True)
        
        if not transferencias:
            lista_trans.controls.append(
                ft.Container(
                    content=ft.Text("No hay transferencias registradas.", italic=True, color=colores["texto_secundario"]),
                    padding=40,
                    alignment=ft.alignment.center
                )
            )
        else:
            for trans in transferencias:
                id_t, origen, destino, monto, fecha, desc = trans
                lista_trans.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Icon("swap_horiz", color=colores["cyan"]),
                            ft.Column([
                                ft.Text(f"{origen} → {destino}", weight=ft.FontWeight.BOLD, size=14, color=colores["texto"]),
                                ft.Text(fecha, size=12, color=colores["texto_secundario"]),
                            ], expand=True, spacing=2),
                            ft.Text(f"${monto:,.0f}", weight=ft.FontWeight.BOLD, color=colores["cyan"], size=16)
                        ]),
                        padding=12,
                        bgcolor=colores["tarjeta"],
                        border_radius=10,
                        border=ft.border.all(1, colores["borde"])
                    )
                )
        
        return ft.Column([
            header,
            ft.Container(
                content=ft.Text("📜 Historial de Transferencias", size=16, weight=ft.FontWeight.BOLD, color=colores["texto"]),
                padding=ft.padding.only(left=15, top=10)
            ),
            lista_trans
        ], spacing=0, expand=True)
    
    # Actualizar función de cambio de vista
    def cambiar_vista(e):
        nonlocal vista_actual
        vistas = ["inicio", "suscripciones", "prestamos", "creditos", "ahorros", "bancos", "balance", "presupuestos", "transferencias", "configuracion"]
        vista_actual = vistas[e.control.selected_index]
        actualizar_vista()
    
    def actualizar_vista():
        """Actualiza la vista actual"""
        nonlocal vista_actual
        
        actualizar_balance()
        contenedor_principal.controls.clear()
        
        if vista_actual == "inicio":
            contenedor_principal.controls.append(crear_vista_inicio())
        elif vista_actual == "suscripciones":
            contenedor_principal.controls.append(crear_vista_suscripciones())
        elif vista_actual == "prestamos":
            contenedor_principal.controls.append(crear_vista_prestamos())
        elif vista_actual == "creditos":
            contenedor_principal.controls.append(crear_vista_creditos())
        elif vista_actual == "ahorros":
            contenedor_principal.controls.append(crear_vista_ahorros())
        elif vista_actual == "bancos":
            contenedor_principal.controls.append(crear_vista_bancos())
        elif vista_actual == "balance":
            contenedor_principal.controls.append(crear_vista_balance_mensual())
        elif vista_actual == "presupuestos":
            contenedor_principal.controls.append(crear_vista_presupuestos())
        elif vista_actual == "transferencias":
            contenedor_principal.controls.append(crear_vista_transferencias())
        elif vista_actual == "configuracion":
            contenedor_principal.controls.append(crear_vista_configuracion())
        
        page.update()

    # Barra superior con botón de configuración
    page.appbar = ft.AppBar(
        title=ft.Text("💰 Mis Finanzas", color="white", size=20),
        center_title=True,
        bgcolor=colores["appbar"],
        elevation=2,
        actions=[
            ft.IconButton(
                icon="settings",
                icon_color="white",
                tooltip="Configuración",
                on_click=lambda e: ir_a_configuracion()
            )
        ]
    )
    
    def ir_a_configuracion():
        nonlocal vista_actual
        vista_actual = "configuracion"
        page.navigation_bar.selected_index = None
        actualizar_vista()

    # Barra de navegación inferior compacta (solo iconos para móvil)
    page.navigation_bar = ft.NavigationBar(
        destinations=[
            ft.NavigationBarDestination(icon="home", label=""),
            ft.NavigationBarDestination(icon="subscriptions", label=""),
            ft.NavigationBarDestination(icon="account_balance", label=""),
            ft.NavigationBarDestination(icon="credit_card", label=""),
            ft.NavigationBarDestination(icon="savings", label=""),
            ft.NavigationBarDestination(icon="account_balance_wallet", label=""),
            ft.NavigationBarDestination(icon="bar_chart", label=""),
            ft.NavigationBarDestination(icon="pie_chart", label=""),
            ft.NavigationBarDestination(icon="swap_horiz", label=""),
        ],
        on_change=cambiar_vista,
        selected_index=0,
        label_behavior=ft.NavigationBarLabelBehavior.ALWAYS_HIDE,
        height=56
    )

    # Botón Flotante
    page.floating_action_button = ft.FloatingActionButton(
        icon="add",
        bgcolor=colores["appbar"],
        on_click=abrir_agregar
    )

    # Agregar BottomSheets al overlay
    page.overlay.append(bottom_sheet_movimiento)
    page.overlay.append(bottom_sheet_suscripcion)
    page.overlay.append(bottom_sheet_prestamo)
    page.overlay.append(bottom_sheet_pago_prestamo)
    page.overlay.append(bottom_sheet_credito)
    page.overlay.append(bottom_sheet_ahorro)
    page.overlay.append(bottom_sheet_monto_ahorro)
    page.overlay.append(bottom_sheet_banco)
    page.overlay.append(bottom_sheet_monto_banco)
    page.overlay.append(dialogo_confirmacion)
    
    # =====================================================
    # PANTALLA DE LOGIN (PIN)
    # =====================================================
    
    contenedor_login.content = ft.Container(
        content=ft.Column([
            ft.Container(height=80),
            ft.Icon("lock", size=80, color="blue700"),
            ft.Container(height=20),
            txt_pin_titulo,
            ft.Text("Ingresa 4 dígitos" if not db.tiene_pin() else "", size=14, color="grey600"),
            ft.Container(height=30),
            ft.Row(pin_inputs, alignment=ft.MainAxisAlignment.CENTER, spacing=10),
            ft.Container(height=10),
            txt_pin_mensaje,
            ft.Container(height=30),
            btn_saltar_pin,
        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
        padding=30,
        expand=True,
        bgcolor=colores["fondo"]
    )
    
    # Configurar título según si hay PIN
    if db.tiene_pin():
        txt_pin_titulo.value = "Ingresa tu PIN"
    else:
        txt_pin_titulo.value = "Crea tu PIN de seguridad"
    
    # Agregar contenedor de la app
    contenedor_app.content = contenedor_principal
    
    # Agregar todo a la página
    page.add(contenedor_login)
    page.add(contenedor_onboarding)
    page.add(contenedor_app)
    
    # Focus en el primer campo de PIN
    try:
        pin_inputs[0].focus()
    except:
        pass
    page.update()

# Ejecutar la app
if __name__ == "__main__":
    ft.app(target=main)