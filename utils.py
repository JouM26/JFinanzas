# utils.py - Funciones utilitarias para JFinanzas
import os
import pathlib
import platform

# Importar openpyxl solo si está disponible (no funciona en Android)
EXCEL_DISPONIBLE = False
Workbook = None
Font = None
PatternFill = None
Alignment = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_DISPONIBLE = True
except:
    pass


def es_android():
    """Detecta si estamos en Android"""
    try:
        # En Android, el home suele contener /data/data/ o /data/user/
        home = str(pathlib.Path.home())
        if '/data/data/' in home or '/data/user/' in home:
            return True
        # También verificar variable de entorno
        if os.environ.get('ANDROID_ROOT') or os.environ.get('ANDROID_DATA'):
            return True
        return False
    except:
        return False


def get_persistent_db_path():
    """
    Obtiene una ruta persistente para la base de datos.
    Esta ruta NO se elimina cuando se actualiza la app.
    """
    try:
        # En Android, usar directorio actual
        if es_android():
            return "finanzas.db"
        
        sistema = platform.system().lower()
        
        if sistema == 'windows':
            app_data_dir = pathlib.Path.home() / "AppData" / "Local" / "JFinanzas"
        elif sistema == 'darwin':
            app_data_dir = pathlib.Path.home() / "Library" / "Application Support" / "JFinanzas"
        elif sistema == 'linux':
            app_data_dir = pathlib.Path.home() / ".jfinanzas"
        else:
            app_data_dir = pathlib.Path.home() / ".jfinanzas"
        
        app_data_dir.mkdir(parents=True, exist_ok=True)
        return str(app_data_dir / "finanzas.db")
    except Exception as e:
        print(f"Error obteniendo ruta persistente: {e}")
        return "finanzas.db"


def obtener_colores(es_oscuro):
    """Retorna diccionario de colores según el tema"""
    if es_oscuro:
        return {
            "fondo": "#0d1117",
            "tarjeta": "#161b22",
            "tarjeta_elevada": "#21262d",
            "texto": "#e6edf3",
            "texto_secundario": "#8b949e",
            "borde": "#30363d",
            "appbar": "#161b22",
            "input_border": "#30363d",
            "input_bg": "#0d1117",
            "verde": "#238636",
            "verde_bg": "#0d1117",
            "rojo": "#da3633",
            "rojo_bg": "#0d1117",
            "naranja": "#d29922",
            "naranja_bg": "#1c1504",
            "purple": "#8957e5",
            "purple_bg": "#1a0d2e",
            "azul": "#58a6ff",
            "azul_bg": "#0d1117",
            "teal": "#3fb950",
            "teal_bg": "#0d1a0f",
            "cyan": "#39c5cf",
            "cyan_bg": "#0a1a1c",
            "indigo": "#a371f7",
            "indigo_bg": "#170d2e",
            "gris_bg": "#21262d",
        }
    else:
        return {
            "fondo": "white",
            "tarjeta": "white",
            "tarjeta_elevada": "#f6f8fa",
            "texto": "black",
            "texto_secundario": "grey600",
            "borde": "#e0e0e0",
            "appbar": "blue700",
            "input_border": "blue700",
            "input_bg": "white",
            "verde": "green",
            "verde_bg": "green50",
            "rojo": "red",
            "rojo_bg": "red50",
            "naranja": "orange",
            "naranja_bg": "orange50",
            "purple": "purple",
            "purple_bg": "purple50",
            "azul": "blue700",
            "azul_bg": "blue50",
            "teal": "teal",
            "teal_bg": "teal50",
            "cyan": "cyan900",
            "cyan_bg": "cyan50",
            "indigo": "indigo",
            "indigo_bg": "indigo50",
            "gris_bg": "grey100",
        }


def exportar_movimientos_a_excel(db, mes, anio):
    """Exporta los movimientos mensuales a un archivo Excel"""
    if not EXCEL_DISPONIBLE:
        return False, "Exportación Excel no disponible en este dispositivo"
    
    try:
        movimientos = db.obtener_movimientos_mensuales(mes, anio)
        ingresos_mes, gastos_mes = db.obtener_balance_mensual(mes, anio)
        total_subs = db.obtener_total_suscripciones()
        total_cuotas = db.obtener_total_cuotas_prestamos()
        total_cuotas_creditos = db.obtener_total_cuotas_creditos()
        balance_mes = ingresos_mes - gastos_mes
        
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        mes_nombre = meses[mes - 1]
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{mes_nombre} {anio}"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Reporte de Movimientos - {mes_nombre} {anio}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A3'] = "RESUMEN DEL MES"
        ws['A3'].font = Font(bold=True, size=12)
        
        ws['A4'] = "Total Ingresos:"
        ws['B4'] = f"${ingresos_mes:,.2f}"
        ws['B4'].font = Font(color="008000", bold=True)
        
        ws['A5'] = "Total Gastos:"
        ws['B5'] = f"${gastos_mes:,.2f}"
        ws['B5'].font = Font(color="FF0000", bold=True)
        
        ws['A6'] = "Suscripciones:"
        ws['B6'] = f"${total_subs:,.2f}"
        
        ws['A7'] = "Cuotas Préstamos:"
        ws['B7'] = f"${total_cuotas:,.2f}"
        
        ws['A8'] = "Cuotas Créditos:"
        ws['B8'] = f"${total_cuotas_creditos:,.2f}"
        
        ws['A9'] = "Balance Final:"
        ws['B9'] = f"${balance_mes:,.2f}"
        ws['B9'].font = Font(bold=True, size=12)
        
        ws['A11'] = "Fecha"
        ws['B11'] = "Tipo"
        ws['C11'] = "Categoría"
        ws['D11'] = "Descripción"
        ws['E11'] = "Monto"
        
        for cell in ['A11', 'B11', 'C11', 'D11', 'E11']:
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = Alignment(horizontal='center')
        
        fila = 12
        for mov in movimientos:
            id_mov, tipo, categoria, monto, descripcion, fecha = mov
            
            ws[f'A{fila}'] = fecha
            ws[f'B{fila}'] = tipo.upper()
            ws[f'C{fila}'] = categoria
            ws[f'D{fila}'] = descripcion
            ws[f'E{fila}'] = f"${monto:,.2f}"
            
            if tipo == 'ingreso':
                ws[f'B{fila}'].font = Font(color="008000")
                ws[f'E{fila}'].font = Font(color="008000")
            else:
                ws[f'B{fila}'].font = Font(color="FF0000")
                ws[f'E{fila}'].font = Font(color="FF0000")
            
            fila += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        
        nombre_archivo = f"Movimientos_{mes_nombre}_{anio}.xlsx"
        ruta_documentos = os.path.join(os.path.expanduser('~'), 'Documents')
        ruta_completa = os.path.join(ruta_documentos, nombre_archivo)
        
        wb.save(ruta_completa)
        return True, ruta_completa
        
    except Exception as e:
        print(f"Error al exportar a Excel: {e}")
        return False, str(e)


# Constantes de categorías
CATEGORIAS = ["Comida", "Transporte", "Servicios", "Ocio", "Salud", "Salario", "Compras", "Educación", "Otro"]

COLORES_CATEGORIAS = {
    "Comida": "#FF6384", 
    "Transporte": "#36A2EB", 
    "Servicios": "#FFCE56",
    "Ocio": "#4BC0C0", 
    "Salud": "#9966FF", 
    "Salario": "#2ECC71",
    "Compras": "#FF9F40", 
    "Educación": "#E74C3C", 
    "Otro": "#95A5A6"
}

MESES_NOMBRES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

MESES_CORTOS = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


# Páginas de onboarding
ONBOARDING_PAGES = [
    {
        "icono": "account_balance_wallet",
        "titulo": "¡Bienvenido a Mis Finanzas!",
        "descripcion": "Tu asistente personal para controlar ingresos, gastos y alcanzar tus metas financieras.",
        "color": "blue"
    },
    {
        "icono": "trending_up",
        "titulo": "Controla tus Movimientos",
        "descripcion": "Registra fácilmente todos tus ingresos y gastos. Categoriza y mantén un historial completo.",
        "color": "green"
    },
    {
        "icono": "subscriptions",
        "titulo": "Gestiona Suscripciones",
        "descripcion": "Nunca pierdas de vista tus pagos recurrentes como Netflix, Spotify y más.",
        "color": "orange"
    },
    {
        "icono": "savings",
        "titulo": "Alcanza tus Metas",
        "descripcion": "Crea metas de ahorro y visualiza tu progreso. ¡Cada peso cuenta!",
        "color": "teal"
    },
    {
        "icono": "pie_chart",
        "titulo": "Visualiza tus Finanzas",
        "descripcion": "Gráficos y reportes para entender mejor cómo gastas tu dinero.",
        "color": "purple"
    }
]
